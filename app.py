
# ============================================================
# QHSE Manager Pro
# Version: v1.0 STABLE
# Copyright © 2026 Kodjotse Eli ADIGBLI. Tous droits réservés.
# ============================================================

from flask import Flask, request, redirect, url_for, session, send_file, render_template_string, flash, get_flashed_messages
import os
import sqlite3
import shutil
from pathlib import Path
from datetime import datetime, date, timedelta
import webbrowser
import threading
import smtplib
import urllib.parse
from email.message import EmailMessage
from markupsafe import escape
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

APP_NAME = "QHSE Manager Pro"
VERSION = "v2.8 ACTIONS + MESSAGERIE + MENUS"
COPYRIGHT = "Copyright © 2026 Kodjotse Eli ADIGBLI. Tous droits réservés."

BASE_DIR = Path(__file__).parent
# Sur Render Starter, DATA_DIR doit pointer vers le disque persistant monté sur /var/data.
# En local, l'application continue d'utiliser le dossier du projet.
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "qhse_chemical_register.db"
EXPORT_DIR = DATA_DIR / "exports"
UPLOAD_DIR = DATA_DIR / "uploads"
BACKUP_DIR = DATA_DIR / "backups"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "qhse_manager_pro_change_this_key")
SESSION_TIMEOUT_MINUTES = max(5, int(os.environ.get("SESSION_TIMEOUT_MINUTES", "30")))
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=SESSION_TIMEOUT_MINUTES)
)
if os.environ.get("CLOUD_MODE", "false").lower() == "true":
    app.config["SESSION_COOKIE_SECURE"] = True

@app.after_request
def add_security_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    return response

@app.before_request
def enforce_session_timeout_and_run_backup():
    """Déconnecte une session inactive et lance les sauvegardes automatiques."""
    now_ts = datetime.now().timestamp()
    if logged():
        last_activity = session.get("_last_activity")
        if last_activity and now_ts - float(last_activity) > SESSION_TIMEOUT_MINUTES * 60:
            session.clear()
            flash(f"Session expirée après {SESSION_TIMEOUT_MINUTES} minutes d’inactivité. Reconnecte-toi.")
            return redirect(url_for("login"))
        session.permanent = True
        session["_last_activity"] = now_ts
        if can_access_admin_features():
            automatic_daily_backup()

CSS = """
<style>
:root{--main:#1f3347;--light:#eef3f7;--border:#cfd8e3;--danger:#b00020;--ok:#087a2c;--warn:#b36b00}
body{font-family:Arial,sans-serif;margin:0;background:#f6f8fb;color:#222}
header{background:var(--main);color:white;padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
header h2{margin:0;font-size:21px}header small{opacity:.9}
nav{background:#e7edf3;padding:10px 24px;border-bottom:1px solid var(--border)}
nav a{margin-right:14px;color:var(--main);font-weight:bold;text-decoration:none}
main{padding:24px}.footer{margin-top:28px;color:#666;font-size:12px}
table{border-collapse:collapse;width:100%;background:white;margin-top:16px}
th,td{border:1px solid var(--border);padding:8px;text-align:left;vertical-align:top}
th{background:var(--main);color:white}
input,select,textarea{width:100%;padding:8px;margin:4px 0 10px 0;box-sizing:border-box;border:1px solid var(--border);border-radius:4px}
button,.btn{background:var(--main);color:white;padding:9px 14px;border:0;border-radius:4px;text-decoration:none;display:inline-block;cursor:pointer}
.box{background:white;padding:16px;border:1px solid var(--border);border-radius:8px;max-width:1200px}
.grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.row{display:grid;grid-template-columns:1fr 1fr;gap:16px}.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px}
.card{background:white;border:1px solid var(--border);border-radius:8px;padding:16px}.value{font-size:28px;font-weight:bold;margin-top:8px}
.danger{color:var(--danger);font-weight:bold}.ok{color:var(--ok);font-weight:bold}.warn{color:var(--warn);font-weight:bold}
.btn-danger{background:var(--danger)}.btn-warn{background:var(--warn)}.actions{white-space:nowrap}.muted{color:#666;font-size:12px}.mini{font-size:12px;padding:5px 8px}
.flash{background:#fff2cc;padding:10px;border:1px solid #e0c36c;margin-bottom:12px;border-radius:6px}
.help{background:#eef7ff;border:1px solid #bed7ee;padding:12px;border-radius:8px;margin-bottom:14px}
.badge{display:inline-block;padding:3px 8px;border-radius:12px;background:#eef3f7;font-size:12px}
.login-wrap{min-height:calc(100vh - 170px);display:flex;align-items:center;justify-content:center;padding:24px}
.login-card{width:100%;max-width:480px;background:white;border:1px solid var(--border);border-radius:12px;padding:24px;box-shadow:0 10px 28px rgba(31,51,71,.14)}
.login-card h2{text-align:center;margin-top:0;color:var(--main)}
.login-card button,.btn-full{width:100%;font-weight:bold}
.login-subtitle{text-align:center;color:#666;font-size:13px;margin-top:-8px;margin-bottom:18px}
@media(max-width:900px){
.grid,.row,.row3{grid-template-columns:1fr}
header{padding:12px 14px;align-items:flex-start;gap:8px;flex-wrap:wrap}header h2{font-size:18px}
nav{padding:8px 12px;white-space:nowrap;overflow-x:auto;-webkit-overflow-scrolling:touch}
nav a{display:inline-block;margin:0 12px 4px 0}
main{padding:12px}.box,.card{padding:12px}.login-wrap{min-height:auto;padding:12px}
table{display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;white-space:nowrap;font-size:13px}
th,td{padding:7px}.btn,button{min-height:40px}.actions{white-space:nowrap}
}
.pagination{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin:14px 0}.pagination a,.pagination span{padding:7px 10px;border:1px solid var(--border);border-radius:4px;text-decoration:none;background:white;color:var(--main)}.pagination .current{background:var(--main);color:white}.pagination select{width:auto;margin:0;padding:6px}
.table-wrap{max-height:68vh;overflow:auto;border:1px solid var(--border);background:white}.table-wrap table{margin-top:0}.table-wrap thead th,.sticky-head th{position:sticky;top:0;z-index:5;background:var(--main)}
.nav-menu{display:inline-block;position:relative;margin-right:8px}.nav-menu>summary{display:inline-block;cursor:pointer;font-weight:bold;color:var(--main);padding:2px 0;list-style:none}.nav-menu>summary::-webkit-details-marker{display:none}.nav-menu[open]>summary{opacity:.75}.nav-drop{position:absolute;z-index:30;min-width:230px;background:white;border:1px solid var(--border);border-radius:6px;box-shadow:0 8px 24px rgba(0,0,0,.14);padding:8px;top:26px;left:0}.nav-drop a{display:block;margin:0;padding:8px;border-radius:4px}.nav-drop a:hover{background:var(--light)}
.priority-critique{background:#ff4d4d!important;color:white;font-weight:bold}.priority-elevee{background:#ff0000!important;color:white;font-weight:bold}.priority-moyenne{background:#ffc000!important;color:#111;font-weight:bold}.priority-faible{background:#92d050!important;color:#111;font-weight:bold}.status-ouvert{background:#92d050!important;color:#111;font-weight:bold}.status-encours{background:#ffc000!important;color:#111;font-weight:bold}.status-ferme{background:#19aee4!important;color:#111;font-weight:bold}.status-suspendu{background:#bfbfbf!important;color:#111;font-weight:bold}
@media(max-width:900px){.nav-menu{position:static}.nav-drop{position:fixed;left:12px;right:12px;top:auto;min-width:auto;max-height:65vh;overflow:auto}.table-wrap{max-height:60vh}}
</style>
"""

# ---------- DB helpers ----------
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def create_db_backup(reason="manual"):
    """Crée une copie sûre de la base SQLite sans interrompre l'application."""
    if not DB_PATH.exists() or DB_PATH.stat().st_size == 0:
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_reason = "".join(ch for ch in reason.lower().replace(" ", "_") if ch.isalnum() or ch in ["_", "-"])[:40] or "backup"
    backup_path = BACKUP_DIR / f"qhse_chemical_register_{safe_reason}_{timestamp}.db"
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(backup_path)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()
    return backup_path

def backup_before_sensitive_operation(reason):
    try:
        path = create_db_backup(reason)
        if path:
            log_action("Sauvegarde automatique", path.name)
        return path
    except Exception as exc:
        print("Backup warning:", exc)
        return None

def latest_backup_file():
    backups = sorted(BACKUP_DIR.glob("qhse_chemical_register_*.db"), key=lambda x: x.stat().st_mtime, reverse=True)
    return backups[0] if backups else None

def automatic_daily_backup():
    """Sauvegarde automatique quotidienne au premier accès de la journée."""
    try:
        today_marker = BACKUP_DIR / f"auto_done_{date.today().isoformat()}.txt"
        if not today_marker.exists() and DB_PATH.exists() and DB_PATH.stat().st_size > 0:
            path = create_db_backup("auto_journalier")
            today_marker.write_text(path.name if path else "no_db", encoding="utf-8")
            if path:
                log_action("Sauvegarde automatique journalière", path.name)
    except Exception as exc:
        print("Daily backup warning:", exc)

def protect_database_on_startup():
    """Protection au démarrage : sauvegarde préventive de la base existante."""
    try:
        marker = BACKUP_DIR / f"startup_done_{date.today().isoformat()}.txt"
        if DB_PATH.exists() and DB_PATH.stat().st_size > 0 and not marker.exists():
            path = create_db_backup("avant_demarrage")
            marker.write_text(path.name if path else "no_db", encoding="utf-8")
    except Exception as exc:
        print("Startup backup warning:", exc)

def validate_backup_filename(filename):
    name = os.path.basename(filename or "")
    path = BACKUP_DIR / name
    if not name.startswith("qhse_chemical_register_") or path.suffix != ".db" or not path.exists():
        return None
    return path

def ensure_column(conn, table, col, definition):
    cols = [r["name"] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    if col not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")

def init_db():
    conn = db()
    cur = conn.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'super_admin',
        subcontractor_id INTEGER,
        project_id INTEGER,
        status TEXT DEFAULT 'Actif'
    )""")
    for col, definition in [
        ("subcontractor_id","INTEGER"),("project_id","INTEGER"),("status","TEXT DEFAULT 'Actif'")
    ]:
        ensure_column(conn, "users", col, definition)

    cur.execute("""CREATE TABLE IF NOT EXISTS projects(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        location TEXT,
        status TEXT DEFAULT 'Actif'
    )""")
    for col, definition in [
        ("client","TEXT"),("owner","TEXT"),("main_contractor","TEXT"),("project_manager","TEXT"),
        ("qhse_manager","TEXT"),("coordinator","TEXT"),("country","TEXT"),("start_date","TEXT"),
        ("end_date","TEXT"),("description","TEXT")
    ]:
        ensure_column(conn, "projects", col, definition)

    cur.execute("""CREATE TABLE IF NOT EXISTS subcontractors(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        contact TEXT,
        phone TEXT,
        email TEXT,
        status TEXT DEFAULT 'Actif'
    )""")
    for col, definition in [
        ("address","TEXT"),("lot","TEXT"),("work_zone","TEXT"),("employees_count","INTEGER DEFAULT 0"),
        ("start_date","TEXT"),("end_date","TEXT"),("hse_plan","TEXT DEFAULT 'NON'"),
        ("insurance","TEXT DEFAULT 'NON'"),("ppsps","TEXT DEFAULT 'NON'"),("observations","TEXT")
    ]:
        ensure_column(conn, "subcontractors", col, definition)

    cur.execute("""CREATE TABLE IF NOT EXISTS products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        root_name TEXT UNIQUE NOT NULL,
        manufacturer TEXT,
        family TEXT,
        physical_state TEXT,
        conditionnement TEXT,
        utilisation TEXT,
        fds TEXT DEFAULT 'NON',
        pictogrammes TEXT,
        incompatibilites TEXT,
        danger_class TEXT,
        epi TEXT,
        observations TEXT
    )""")
    for col, definition in [
        ("commercial_name","TEXT"),("reference","TEXT"),("h_statements","TEXT"),("p_statements","TEXT"),
        ("first_aid","TEXT"),("fire_measures","TEXT"),("spill_measures","TEXT"),("storage_rules","TEXT"),
        ("fds_date","TEXT"),("fds_version","TEXT"),("fds_expiry","TEXT"),
        ("stock_min","REAL DEFAULT 0"),("stock_max","REAL DEFAULT 0")
    ]:
        ensure_column(conn, "products", col, definition)

    cur.execute("""CREATE TABLE IF NOT EXISTS aliases(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alias_name TEXT UNIQUE NOT NULL,
        product_id INTEGER NOT NULL
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER,
        entry_date TEXT NOT NULL,
        month TEXT NOT NULL,
        subcontractor_id INTEGER NOT NULL,
        declared_product TEXT NOT NULL,
        product_id INTEGER,
        qty_in REAL DEFAULT 0,
        qty_used REAL DEFAULT 0,
        qty_stock REAL DEFAULT 0,
        unit TEXT DEFAULT 'u',
        storage_location TEXT,
        fds_available TEXT DEFAULT 'NON',
        observations TEXT
    )""")
    ensure_column(conn, "entries", "utilisation", "TEXT")


    cur.execute("""CREATE TABLE IF NOT EXISTS daily_reports(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_date TEXT NOT NULL,
        project_id INTEGER NOT NULL,
        subcontractor_id INTEGER NOT NULL,
        status TEXT DEFAULT 'Brouillon',
        general_observations TEXT,
        created_by TEXT,
        created_at TEXT,
        updated_at TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS daily_report_fields(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        label TEXT NOT NULL,
        unit TEXT,
        display_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS daily_report_columns(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        display_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS daily_report_values(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_id INTEGER NOT NULL,
        field_id INTEGER NOT NULL,
        column_id INTEGER NOT NULL,
        value TEXT,
        UNIQUE(report_id, field_id, column_id)
    )""")

    ensure_column(conn, "projects", "whatsapp_group_link", "TEXT")
    ensure_column(conn, "projects", "notification_email", "TEXT")

    cur.execute("""CREATE TABLE IF NOT EXISTS action_plans(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_id INTEGER,
        source TEXT,
        opened_date TEXT NOT NULL,
        risk_description TEXT NOT NULL,
        required_action TEXT,
        responsible TEXT,
        due_date TEXT,
        priority TEXT DEFAULT 'Moyenne',
        status TEXT DEFAULT 'Ouvert',
        comments TEXT,
        created_by TEXT,
        created_at TEXT,
        updated_at TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS message_log(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sent_at TEXT,
        sender TEXT,
        channel TEXT,
        subject TEXT,
        message TEXT,
        recipients TEXT,
        project_id INTEGER,
        subcontractor_id INTEGER,
        delivery_status TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS audit_log(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action_date TEXT,
        username TEXT,
        action TEXT,
        details TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS workforce(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_date TEXT NOT NULL,
        project_id INTEGER,
        subcontractor_id INTEGER NOT NULL,
        men INTEGER DEFAULT 0,
        women INTEGER DEFAULT 0,
        total INTEGER DEFAULT 0,
        observations TEXT,
        created_by TEXT,
        created_at TEXT
    )""")

    cur.execute("UPDATE users SET role='super_admin' WHERE role='admin'")
    if cur.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] == 0:
        admin_username = os.environ.get("ADMIN_USERNAME", "admin")
        admin_password = os.environ.get("ADMIN_PASSWORD")
        cloud_mode = os.environ.get("CLOUD_MODE", "false").lower() == "true"
        if not admin_password:
            if cloud_mode:
                raise RuntimeError("ADMIN_PASSWORD doit être défini dans Render.")
            admin_password = "admin123"
        cur.execute("INSERT INTO users(username,password_hash,role,status) VALUES(?,?,?,?)",
                    (admin_username, generate_password_hash(admin_password), "super_admin", "Actif"))

    cur.execute("INSERT OR IGNORE INTO projects(name,location,country,main_contractor,status) VALUES(?,?,?,?,?)",
                ("CHU Kara", "Kara", "Togo", "Ellipse Projects", "Actif"))
    cur.execute("INSERT OR IGNORE INTO projects(name,location,country,main_contractor,status) VALUES(?,?,?,?,?)",
                ("CHU Campus Lomé", "Lomé", "Togo", "Ellipse Projects", "Actif"))

    for st in ["SICONE","CENTRO","BETEIR","BATICO","MGP"]:
        cur.execute("INSERT OR IGNORE INTO subcontractors(name,status) VALUES(?,?)", (st, "Actif"))

    sample_products = [
        ("AQUADERE","Résine / primaire","Liquide","Bidon / seau","Résine d’accrochage pour mortier et béton","NON","","","Résine / primaire","Gants, lunettes",0,0),
        ("SOPRALENE","Étanchéité","Solide","Rouleau membrane","Étanchéité des toitures et ouvrages","NON","","Éloigner sources de chaleur","Produit bitumineux","Gants, chaussures de sécurité",5,250),
        ("EQUERRE DE KENF","Accessoire","Solide","Carton / pièce","Accessoire de fixation et assemblage","NON","","","Accessoire","Gants",0,0),
        ("PANTEX VELOUR","Peinture","Liquide","Seau plastique 15 à 25 kg","Peinture de finition intérieure murs et plafonds","NON","SGH07 possible","Oxydants forts","Peinture","Gants, lunettes",5,300),
        ("PRIMAIRE D ACCROCHE","Primaire","Liquide","Seau 5 à 20 kg","Préparation des supports avant revêtement","NON","SGH07 possible","Oxydants forts","Primaire","Gants, lunettes",0,200),
        ("SIGMA COVER","Peinture","Liquide","Seau peinture","Peinture de finition et protection","NON","SGH07 possible","Oxydants forts","Peinture","Gants, lunettes",0,100),
        ("FLINTKOTE IKOPRIM","Étanchéité","Liquide","Bidon / seau","Primaire bitumineux pour travaux d’étanchéité","NON","SGH02/SGH07 possible","Sources d’ignition, oxydants","Bitume / primaire","Gants, lunettes, ventilation",0,80),
        ("FLINTKOTE BE6","Étanchéité","Liquide","Seau / bidon","Revêtement bitumineux de protection et d’étanchéité","NON","SGH02/SGH07 possible","Sources d’ignition, oxydants","Bitume","Gants, lunettes",0,80),
        ("ULTRABONDE CONDUCTIVITE","Colle","Pâte","Seau 16 kg","Colle conductrice pour revêtements de sol techniques","NON","SGH07 possible","Oxydants forts","Colle","Gants, lunettes",0,80),
        ("ULTRABONDE EVOLUTION","Colle","Pâte","Seau 14 kg","Colle pour revêtements de sols et murs","NON","SGH07 possible","Oxydants forts","Colle","Gants, lunettes",0,150),
        ("MC TECHNIFLOW 57GH","Mortier / Coulis","Poudre","Sac 25 kg","Coulis de scellement et calage haute performance","NON","SGH07 possible","Acides forts","Coulis cimentaire","Gants, lunettes, masque poussières",0,100),
        ("AFRICA ENDUIT","Enduit","Poudre","Sac 25 kg","Enduit de lissage et de finition des murs","NON","SGH07 possible","Acides forts","Enduit","Gants, masque poussières",0,150),
        ("POT DE PLATRE SAFEBOARD","Plâtre","Poudre","Sac / seau","Traitement des joints et finition des plaques de plâtre","NON","SGH07 possible","Humidité excessive","Plâtre","Gants, masque poussières",0,50),
        ("DILUAN DJESSI","Solvant","Liquide","Bidon","Dilution des peintures et nettoyage du matériel","NON","SGH02/SGH07 possible","Sources d’ignition, oxydants","Diluant","Gants, lunettes, ventilation",0,100),
        ("LAC PRO","Vernis","Liquide","Bidon / seau","Vernis et protection des surfaces peintes","NON","SGH02/SGH07 possible","Sources d’ignition, oxydants","Vernis","Gants, lunettes",0,50),
        ("PEINTURE DJESSI","Peinture","Liquide","Seau peinture","Peinture intérieure et extérieure pour bâtiment","NON","SGH07 possible","Oxydants forts","Peinture","Gants, lunettes",0,150),
        ("SIKA MORTIER","Mortier","Poudre","Sac 25 kg","Réparation, scellement et reprofilage du béton","NON","SGH05/SGH07 possible","Acides forts","Mortier cimentaire","Gants, lunettes, masque poussières",0,80),
        ("SIKALASTIC","Étanchéité","Liquide","Seau ou kit","Membrane liquide d’étanchéité","NON","SGH07 possible","Oxydants forts","Étanchéité liquide","Gants, lunettes",0,100),
        ("ALSAN FLASHING","Étanchéité","Liquide","Seau ou kit","Résine liquide pour relevés et détails d’étanchéité","NON","SGH02/SGH07 possible","Sources d’ignition, oxydants","Résine d’étanchéité","Gants, lunettes, ventilation",0,60),
        ("ENDUIT EVOLUTION","Enduit","Poudre","Sac 14 ou 30 kg","Enduit de rebouchage et de lissage","NON","SGH07 possible","Acides forts","Enduit","Gants, masque poussières",0,150),
        ("IMPRIM LUX COLORIS","Primaire","Liquide","Seau","Sous-couche et primaire avant peinture","NON","SGH07 possible","Oxydants forts","Primaire peinture","Gants, lunettes",0,150),
        ("GARNYTEX","Revêtement","Liquide","Seau","Revêtement décoratif et protection des surfaces","NON","SGH07 possible","Oxydants forts","Revêtement décoratif","Gants, lunettes",0,100),
        ("COLLE SPM ACRYLIQUE","Colle","Pâte","Seau / cartouche","Collage de revêtements et matériaux de finition","NON","SGH07 possible","Oxydants forts","Colle acrylique","Gants, lunettes",0,100),
        ("PANTIPRIM","Primaire","Liquide","Seau plastique","Couche d’impression avant peinture ou enduit","NON","SGH07 possible","Oxydants forts","Primaire","Gants, lunettes",0,100),
        ("COLORIS SENIOR","Peinture","Liquide","Seau peinture","Peinture de finition intérieure/extérieure","NON","SGH07 possible","Oxydants forts","Peinture","Gants, lunettes",0,100),
        ("BELLE & REBELLE","Peinture","Liquide","Seau peinture","Peinture décorative de finition","NON","SGH07 possible","Oxydants forts","Peinture décorative","Gants, lunettes",0,80),
        ("SIKA CEM HYDROFUGE LIQUIDE","Adjuvant","Liquide","Bidon","Adjuvant hydrofuge pour mortiers et bétons","NON","SGH07 possible","Oxydants forts","Adjuvant hydrofuge","Gants, lunettes",0,50),
        ("SOLUTION ENDUIT CIMENT","Enduit","Poudre","Sac 25 kg","Enduit ciment pour maçonnerie et façades","NON","SGH05/SGH07 possible","Acides forts","Enduit cimentaire","Gants, lunettes, masque poussières",0,200),
    ]

    for p in sample_products:
        cur.execute("""INSERT OR IGNORE INTO products(root_name,family,physical_state,conditionnement,utilisation,fds,pictogrammes,incompatibilites,danger_class,epi,stock_min,stock_max)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", p)

    if cur.execute("SELECT COUNT(*) c FROM daily_report_columns").fetchone()["c"] == 0:
        cur.execute("INSERT INTO daily_report_columns(name,display_order,active) VALUES('Valeur',1,1)")
        cur.execute("INSERT INTO daily_report_columns(name,display_order,active) VALUES('Observations',2,1)")
    if cur.execute("SELECT COUNT(*) c FROM daily_report_fields").fetchone()["c"] == 0:
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INFORMATIONS ENTREPRISE', 'Type d’entreprise', '', 1))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INFORMATIONS ENTREPRISE', 'Nom de l’entreprise', '', 2))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INFORMATIONS ENTREPRISE', 'Effectif total', 'Nb.', 3))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Réunion hebdomadaire HSE', 'Nb.', 4))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Réunion quart d’heure sécurité (toolbox meeting)', 'Nb.', 5))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Minutes sécurité consécutives à un stop chantier', 'Nb.', 6))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Formation HSE aux travailleurs – joindre copie formation et fiche de présence', 'Nb.', 7))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Nombre d’accueil HSE initial aux travailleurs', 'Nb.', 8))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Inspection HSE – joindre une copie', 'Nb.', 9))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Exercice et formation à la gestion des situations d’urgence', 'Nb.', 10))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS PROACTIFS', 'Campagne HSE', 'Nb.', 11))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS RÉACTIFS', 'Accidents avec arrêt', 'Nb.', 12))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS RÉACTIFS', 'Accidents sans arrêt', 'Nb.', 13))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS RÉACTIFS', 'Incidents / presque-accidents', 'Nb.', 14))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS RÉACTIFS', 'Cas de premiers soins', 'Nb.', 15))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('INDICATEURS RÉACTIFS', 'Décès liés au travail', 'Nb.', 16))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Nombre de recrutements', 'Nb.', 17))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Recrutements de travailleurs issus des listes communautaires', 'Nb.', 18))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Sensibilisation sur l’égalité des sexes, les droits des personnes vulnérables et les risques de contamination par les MST et le VIH/sida', 'Nb.', 19))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Nombre de femmes employées au poste de Manager', 'Nb.', 20))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Nombre de femmes employées', 'Nb.', 21))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA MAIN-D’ŒUVRE', 'Nombre de travailleurs handicapés', 'Nb.', 22))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('PLAN D’INTERVENTION D’URGENCE', 'Équipements opérationnels de lutte contre l’incendie sur le site', 'Nb.', 23))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA CIRCULATION', 'Contrôles d’alcoolémie effectués', 'Nb.', 24))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA CIRCULATION', 'Contrôles des permis / autorisation de conduite d’engin', 'Nb.', 25))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA CIRCULATION', 'NC liée à des manques de panneaux d’information ou signalisation', 'Nb.', 26))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA CIRCULATION', 'Non-conformité liée à l’état général des véhicules et engins de chantier', 'Nb.', 27))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA CIRCULATION', 'Conducteurs d’engins/véhicules sans permis ou autorisations de conduite spécifiques', 'Nb.', 28))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES TRAVAUX DANGEREUX', 'Procédures écrites travaux à risques', 'O/N', 29))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES TRAVAUX DANGEREUX', 'Formation sur la gestion des travaux', 'Nb.', 30))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES TRAVAUX DANGEREUX', 'Formation au système de permis de travail', 'Nb.', 31))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES TRAVAUX DANGEREUX', 'Nombre de permis de travail en cours', 'Nb.', 32))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES TRAVAUX DANGEREUX', 'Nombre d’Analyses des risques préliminaires (ARP)', 'Nb.', 33))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES MATIÈRES DANGEREUSES', 'NC sur le stockage, manipulation, transfert et utilisation des matières dangereuses', 'Nb.', 34))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES MATIÈRES DANGEREUSES', 'Nombre de fiches de données de sécurité', 'Nb.', 35))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES MATIÈRES DANGEREUSES', 'Produits dangereux présents sur site sans FDS', 'Nb.', 36))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('BRUITS ET VIBRATIONS', 'Nombre de mesures de bruit sur le site', 'Nb.', 37))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA BIODIVERSITÉ', 'Individus d’animaux sauvages morts', 'Nb.', 38))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DE LA BIODIVERSITÉ', 'Animaux sauvages rencontrés', 'Nb.', 39))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('PROTECTION DU PATRIMOINE CULTUREL ET ARCHÉOLOGIQUE', 'Travailleurs formés aux découvertes fortuites', 'Nb.', 40))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('PROTECTION DU PATRIMOINE CULTUREL ET ARCHÉOLOGIQUE', 'Nombre de découvertes fortuites', 'Nb.', 41))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES EAUX', 'Consommation d’eau chantier', 'm³', 42))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES DÉCHETS ET EFFLUENTS', 'NC observées sur les équipements chantier', 'Nb.', 43))
        cur.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)", ('GESTION DES DÉCHETS ET EFFLUENTS', 'NC observées sur le tri et la collecte des déchets', 'Nb.', 44))

    conn.commit()
    conn.close()

def log_action(action, details=""):
    try:
        conn = db()
        conn.execute("INSERT INTO audit_log(action_date,username,action,details) VALUES(?,?,?,?)",
                     (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session.get("username","system"), action, details))
        conn.commit()
        conn.close()
    except Exception:
        pass

# ---------- Auth / roles ----------
def role(): return session.get("role")
def is_super_admin(): return role() == "super_admin"
def is_project_admin(): return role() == "project_admin"
def is_subcontractor(): return role() == "subcontractor"
def is_reader(): return role() == "reader"
def current_project_id(): return session.get("project_id")
def current_subcontractor_id(): return session.get("subcontractor_id")
def can_access_admin_features(): return is_super_admin() or is_project_admin()
def logged(): return session.get("user_id") is not None

def can_manage_entry(entry_row):
    if is_super_admin():
        return True
    if is_project_admin():
        return str(entry_row["project_id"]) == str(current_project_id())
    if is_subcontractor():
        return str(entry_row["subcontractor_id"]) == str(current_subcontractor_id())
    return False


def can_manage_workforce(workforce_row):
    if is_super_admin():
        return True
    if is_project_admin():
        return str(workforce_row["project_id"]) == str(current_project_id())
    if is_subcontractor():
        return str(workforce_row["subcontractor_id"]) == str(current_subcontractor_id())
    return False


def can_manage_project(project_row):
    if is_super_admin():
        return True
    if is_project_admin():
        return str(project_row["id"]) == str(current_project_id())
    return False

def can_delete_project(project_row):
    return is_super_admin()

def can_manage_subcontractor(st_row):
    if is_super_admin():
        return True
    if is_project_admin():
        return True
    if is_subcontractor():
        return str(st_row["id"]) == str(current_subcontractor_id())
    return False

def can_delete_subcontractor(st_row):
    return is_super_admin() or is_project_admin()

def can_manage_user(user_row):
    if is_super_admin():
        return True
    if is_project_admin():
        return str(user_row["project_id"]) == str(current_project_id()) and user_row["role"] in ["subcontractor", "reader"]
    return False

def can_delete_user(user_row):
    if str(user_row["id"]) == str(session.get("user_id")):
        return False
    return can_manage_user(user_row)

def yes_no_options(value):
    value = value or "NON"
    return "".join([f"<option {'selected' if value==v else ''}>{v}</option>" for v in ["OUI","NON"]])

def status_options(value):
    value = value or "Actif"
    return "".join([f"<option {'selected' if value==v else ''}>{v}</option>" for v in ["Actif","Inactif","Clôturé"]])

def safe_int(value, default=0):
    try:
        return int(value or default)
    except Exception:
        return default

def safe_float(value, default=0):
    try:
        return float(value or default)
    except Exception:
        return default


def pagination_data(default_per_page=20):
    page = max(safe_int(request.args.get("page"), 1), 1)
    per_page = safe_int(request.args.get("per_page"), default_per_page)
    if per_page != 20:
        per_page = 20
    return page, per_page, (page - 1) * per_page


def pagination_html(page, per_page, total):
    if total <= per_page and page == 1:
        return ""
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    args = request.args.to_dict(flat=True)
    links = []
    if page > 1:
        args.update(page=page - 1, per_page=per_page)
        links.append(f"<a href='{url_for(request.endpoint, **args)}'>← Précédent</a>")
    links.append(f"<span class='current'>Page {page} / {pages}</span>")
    links.append(f"<span>{total} ligne(s)</span>")
    if page < pages:
        args.update(page=page + 1, per_page=per_page)
        links.append(f"<a href='{url_for(request.endpoint, **args)}'>Suivant →</a>")
    return "<div class='pagination'>" + "".join(links) + "</div>"


def where_scope(prefix="e"):
    if is_super_admin():
        return "", []
    if is_project_admin() or is_reader():
        return f" WHERE {prefix}.project_id=? ", [current_project_id()]
    if is_subcontractor():
        return f" WHERE {prefix}.subcontractor_id=? ", [current_subcontractor_id()]
    return " WHERE 1=0 ", []

def workforce_scope(prefix="w"):
    if is_super_admin():
        return "", []
    if is_project_admin() or is_reader():
        return f" WHERE {prefix}.project_id=? ", [current_project_id()]
    if is_subcontractor():
        return f" WHERE {prefix}.subcontractor_id=? ", [current_subcontractor_id()]
    return " WHERE 1=0 ", []

def scoped_and_clause():
    where, params = where_scope("e")
    return (where + (" AND " if where else " WHERE "), params)

# ---------- Product intelligence ----------
def normalize(text):
    text = (text or "").upper()
    for ch in ["-", "_", ".", ",", "/", "\\", "(", ")", "[", "]", "&"]:
        text = text.replace(ch, " ")
    remove = {"KG","KGS","L","LTR","LITRE","LITRES","SEAU","BIDON","SAC","POT","16","14","20","25","30"}
    return " ".join([w for w in text.split() if w not in remove])

def find_product_id(name):
    n = normalize(name)
    if not n:
        return None
    conn = db()
    row = conn.execute("SELECT product_id FROM aliases WHERE alias_name=?", (n,)).fetchone()
    if row:
        conn.close()
        return row["product_id"]

    best_id, best_score = None, 0
    nw = set(n.split())
    for p in conn.execute("SELECT id,root_name FROM products").fetchall():
        r = normalize(p["root_name"])
        rw = set(r.split())
        if n == r:
            conn.close()
            return p["id"]
        if r in n or n in r:
            score = 90
        else:
            score = int(100 * len(nw.intersection(rw)) / max(len(rw), 1))
        if score > best_score:
            best_score, best_id = score, p["id"]
    conn.close()
    return best_id if best_score >= 60 else None

def days_to(date_text):
    try:
        d = datetime.strptime(date_text, "%Y-%m-%d").date()
        return (d - date.today()).days
    except Exception:
        return None

# ---------- UI ----------
def menu_group(label, links):
    items = "".join([f'<a href="{href}">{text}</a>' for text, href in links])
    return f'<details class="nav-menu"><summary>{label} ▾</summary><div class="nav-drop">{items}</div></details>'

def layout(title, body):
    nav = ""
    if logged():
        groups = []
        if can_access_admin_features():
            groups.append(menu_group("Administration", [
                ("Projets", url_for("projects")), ("Utilisateurs", url_for("users")),
                ("Sous-traitants", url_for("subcontractors")), ("Journal des actions", url_for("audit_trail")),
                ("Sauvegardes", url_for("backups"))
            ]))
        groups.append(menu_group("Produits chimiques", [
            ("Saisie des produits", url_for("entries")), ("Base produits", url_for("products")),
            ("Alias / Doublons", url_for("aliases")), ("Alertes", url_for("alerts")),
            ("Consolidation", url_for("consolidation")), ("Export Excel", url_for("export_excel"))
        ]))
        report_links = [("Rapports journaliers", url_for("daily_reports")), ("Effectifs journaliers", url_for("workforce"))]
        if can_access_admin_features():
            report_links += [("Configurer le modèle", url_for("daily_report_template")), ("Export & statistiques", url_for("export_daily_reports_excel")), ("Export effectifs", url_for("export_workforce_excel"))]
        groups.append(menu_group("Rapports journaliers", report_links))
        if can_access_admin_features():
            groups.append(menu_group("Pilotage", [("Plan d’action mensuel", url_for("action_plans"))]))
            groups.append(menu_group("Messagerie", [("Centre de messages", url_for("message_center")), ("Historique des envois", url_for("message_history"))]))
        nav = '<nav><a href="{}">Tableau de bord</a>{}<a href="{}">À propos</a><a href="{}">Déconnexion</a></nav>'.format(
            url_for('dashboard'), ''.join(groups), url_for('about'), url_for('logout'))
    flashes = "".join([f"<div class='flash'>{m}</div>" for m in get_flashed_messages()])
    userlabel = session.get("username","")
    if role(): userlabel += f" | {role()}"
    return render_template_string(f'''<!doctype html>
    <html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>{title}</title>{CSS}</head>
    <body><header><div><h2>{APP_NAME}</h2><small>{VERSION} — {COPYRIGHT}</small></div><small>{userlabel}</small></header>
    {nav}<main>{flashes}{body}<div class="footer">© 2026 Kodjotse Eli ADIGBLI – QHSE Manager Pro</div></main>
    <script>document.addEventListener('click',function(e){{document.querySelectorAll('.nav-menu[open]').forEach(function(d){{if(!d.contains(e.target))d.removeAttribute('open');}});}});</script>
    </body></html>''')

# ---------- Data queries ----------
def consolidation_rows():
    conn = db()
    where, params = where_scope("e")
    rows = conn.execute(f"""SELECT COALESCE(p.root_name,'PRODUIT NON RECONNU: '||e.declared_product) product_name,
    p.family,p.conditionnement,COALESCE(e.utilisation,p.utilisation) utilisation,p.fds,p.fds_expiry,p.pictogrammes,p.incompatibilites,
    p.danger_class,p.epi,p.stock_min,p.stock_max,e.unit,
    SUM(e.qty_in) total_in,SUM(e.qty_used) total_used,SUM(e.qty_stock) total_stock,
    COUNT(DISTINCT e.subcontractor_id) nb_st,
    GROUP_CONCAT(DISTINCT s.name) sts,
    GROUP_CONCAT(DISTINCT pr.name) projects,
    GROUP_CONCAT(DISTINCT e.storage_location) locations
    FROM entries e 
    LEFT JOIN products p ON p.id=e.product_id 
    JOIN subcontractors s ON s.id=e.subcontractor_id 
    LEFT JOIN projects pr ON pr.id=e.project_id
    {where}
    GROUP BY product_name,e.unit 
    ORDER BY product_name""", params).fetchall()
    conn.close()
    return rows

def get_alert_rows():
    alerts = []
    for r in consolidation_rows():
        product = r["product_name"]
        if product.startswith("PRODUIT NON RECONNU"):
            alerts.append(("CRITIQUE","Produit non reconnu",product,"Créer un produit racine ou un alias."))
        if (r["fds"] or "NON") == "NON":
            alerts.append(("CRITIQUE","FDS manquante",product,"Demander la FDS et compléter la base produit."))
        d = days_to(r["fds_expiry"])
        if d is not None and d < 0:
            alerts.append(("CRITIQUE","FDS expirée",product,f"FDS expirée depuis {-d} jour(s)."))
        elif d is not None and d <= 30:
            alerts.append(("ATTENTION","FDS proche expiration",product,f"FDS expire dans {d} jour(s)."))
        stock = r["total_stock"] or 0
        if (r["stock_min"] or 0) > 0 and stock < r["stock_min"]:
            alerts.append(("ATTENTION","Stock critique",product,f"Stock {stock} inférieur au minimum {r['stock_min']}."))
        if (r["stock_max"] or 0) > 0 and stock > r["stock_max"]:
            alerts.append(("CRITIQUE","Stock dépassé",product,f"Stock {stock} supérieur au maximum {r['stock_max']}."))
    return alerts

# ---------- Routes ----------
@app.route("/")
def home():
    return redirect(url_for("dashboard") if logged() else url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        conn = db()
        user = conn.execute("SELECT * FROM users WHERE username=? AND status='Actif'", (request.form["username"],)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], request.form["password"]):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            session["subcontractor_id"] = user["subcontractor_id"]
            session["project_id"] = user["project_id"]
            log_action("Connexion", "Connexion utilisateur")
            return redirect(url_for("dashboard"))
        flash("Identifiants incorrects ou compte inactif.")
    return layout("Connexion", """<div class="login-wrap"><div class="login-card">
        <h2>Connexion</h2><div class="login-subtitle">Accès sécurisé au registre QHSE</div>
        <form method="post" autocomplete="off">
        <label>Utilisateur</label><input name="username" autocomplete="off" required autofocus placeholder="Saisir votre identifiant">
        <label>Mot de passe</label><input name="password" type="password" autocomplete="new-password" required placeholder="Saisir votre mot de passe">
        <button>Se connecter</button></form></div></div>""")

@app.route("/logout")
def logout():
    log_action("Déconnexion", "Déconnexion utilisateur")
    session.clear()
    response = redirect(url_for("login"))
    response.delete_cookie(app.config.get("SESSION_COOKIE_NAME", "session"))
    return response

@app.route("/dashboard")
def dashboard():
    if not logged(): return redirect(url_for("login"))
    conn = db()
    where, params = where_scope("e")
    and_clause, and_params = scoped_and_clause()

    if is_super_admin():
        total_projects = conn.execute("SELECT COUNT(*) c FROM projects").fetchone()["c"]
        total_st = conn.execute("SELECT COUNT(*) c FROM subcontractors").fetchone()["c"]
        total_products = conn.execute("SELECT COUNT(*) c FROM products").fetchone()["c"]
        help_text = "Vue Super Admin : tous les projets, tous les sous-traitants, consolidation générale."
    elif is_project_admin():
        total_projects = 1
        total_st = conn.execute("SELECT COUNT(DISTINCT subcontractor_id) c FROM entries WHERE project_id=?", (current_project_id(),)).fetchone()["c"]
        total_products = conn.execute("SELECT COUNT(DISTINCT product_id) c FROM entries WHERE project_id=? AND product_id IS NOT NULL", (current_project_id(),)).fetchone()["c"]
        help_text = "Vue Admin Projet : tu gères uniquement le projet qui t'est attribué."
    elif is_reader():
        total_projects = 1
        total_st = conn.execute("SELECT COUNT(DISTINCT subcontractor_id) c FROM entries WHERE project_id=?", (current_project_id(),)).fetchone()["c"]
        total_products = conn.execute("SELECT COUNT(DISTINCT product_id) c FROM entries WHERE project_id=? AND product_id IS NOT NULL", (current_project_id(),)).fetchone()["c"]
        help_text = "Vue Lecteur : consultation limitée aux rapports autorisés."
    else:
        total_projects = 1
        total_st = 1
        total_products = conn.execute("SELECT COUNT(DISTINCT declared_product) c FROM entries WHERE subcontractor_id=?", (current_subcontractor_id(),)).fetchone()["c"]
        help_text = "Vue Sous-traitant : tu saisis uniquement tes propres produits chimiques."

    total_stock = conn.execute(f"SELECT COALESCE(SUM(qty_stock),0) s FROM entries e {where}", params).fetchone()["s"]
    total_in = conn.execute(f"SELECT COALESCE(SUM(qty_in),0) s FROM entries e {where}", params).fetchone()["s"]
    total_used = conn.execute(f"SELECT COALESCE(SUM(qty_used),0) s FROM entries e {where}", params).fetchone()["s"]
    unknown = conn.execute(f"SELECT COUNT(*) c FROM entries e {and_clause} e.product_id IS NULL", and_params).fetchone()["c"]
    no_fds = conn.execute(f"SELECT COUNT(*) c FROM entries e {and_clause} e.fds_available='NON'", and_params).fetchone()["c"]

    alert_rows = get_alert_rows()
    alerts_count = len(alert_rows)
    critical_alerts = sum(1 for a in alert_rows if a[0] == "CRITIQUE")
    warning_alerts = sum(1 for a in alert_rows if a[0] == "ATTENTION")

    top_products = conn.execute(f"""SELECT COALESCE(p.root_name,e.declared_product) product, SUM(e.qty_stock) stock
                          FROM entries e LEFT JOIN products p ON p.id=e.product_id
                          {where}
                          GROUP BY product ORDER BY stock DESC LIMIT 10""", params).fetchall()

    top_used = conn.execute(f"""SELECT COALESCE(p.root_name,e.declared_product) product, SUM(e.qty_used) used_qty
                          FROM entries e LEFT JOIN products p ON p.id=e.product_id
                          {where}
                          GROUP BY product ORDER BY used_qty DESC LIMIT 10""", params).fetchall()

    by_subcontractor = conn.execute(f"""SELECT s.name st, SUM(e.qty_stock) stock, COUNT(DISTINCT COALESCE(e.product_id, e.declared_product)) products_count
                          FROM entries e JOIN subcontractors s ON s.id=e.subcontractor_id
                          {where}
                          GROUP BY s.name ORDER BY stock DESC LIMIT 10""", params).fetchall()

    by_family = conn.execute(f"""SELECT COALESCE(p.family,'Non classé') family, COUNT(DISTINCT COALESCE(e.product_id, e.declared_product)) count_products, SUM(e.qty_stock) stock
                          FROM entries e LEFT JOIN products p ON p.id=e.product_id
                          {where}
                          GROUP BY family ORDER BY stock DESC LIMIT 10""", params).fetchall()

    by_month = conn.execute(f"""SELECT e.month, SUM(e.qty_in) qty_in, SUM(e.qty_used) qty_used, SUM(e.qty_stock) stock
                          FROM entries e
                          {where}
                          GROUP BY e.month ORDER BY e.month DESC LIMIT 6""", params).fetchall()

    w_where, w_params = workforce_scope("w")
    today = date.today().strftime("%Y-%m-%d")
    w_and = (w_where + " AND ") if w_where else " WHERE "
    workforce_today = conn.execute(f"SELECT COALESCE(SUM(total),0) s FROM workforce w {w_and} w.work_date=?", w_params + [today]).fetchone()["s"]
    workforce_week = conn.execute(f"SELECT COALESCE(SUM(total),0) s FROM workforce w {w_and} strftime('%Y-%W', w.work_date)=strftime('%Y-%W','now')", w_params).fetchone()["s"]
    workforce_month = conn.execute(f"SELECT COALESCE(SUM(total),0) s FROM workforce w {w_and} substr(w.work_date,1,7)=strftime('%Y-%m','now')", w_params).fetchone()["s"]
    workforce_year = conn.execute(f"SELECT COALESCE(SUM(total),0) s FROM workforce w {w_and} substr(w.work_date,1,4)=strftime('%Y','now')", w_params).fetchone()["s"]
    workforce_by_st = conn.execute(f"""SELECT s.name st, COALESCE(SUM(w.total),0) total
                          FROM workforce w JOIN subcontractors s ON s.id=w.subcontractor_id
                          {w_where}
                          GROUP BY s.name ORDER BY total DESC LIMIT 10""", w_params).fetchall()

    conn.close()

    top_products_rows = "".join([f"<tr><td>{r['product']}</td><td>{r['stock']}</td></tr>" for r in top_products]) or "<tr><td colspan='2'>Aucune donnée</td></tr>"
    top_used_rows = "".join([f"<tr><td>{r['product']}</td><td>{r['used_qty']}</td></tr>" for r in top_used]) or "<tr><td colspan='2'>Aucune donnée</td></tr>"
    st_rows = "".join([f"<tr><td>{r['st']}</td><td>{r['products_count']}</td><td>{r['stock']}</td></tr>" for r in by_subcontractor]) or "<tr><td colspan='3'>Aucune donnée</td></tr>"
    family_rows = "".join([f"<tr><td>{r['family']}</td><td>{r['count_products']}</td><td>{r['stock']}</td></tr>" for r in by_family]) or "<tr><td colspan='3'>Aucune donnée</td></tr>"
    month_rows = "".join([f"<tr><td>{r['month']}</td><td>{r['qty_in']}</td><td>{r['qty_used']}</td><td>{r['stock']}</td></tr>" for r in by_month]) or "<tr><td colspan='4'>Aucune donnée</td></tr>"
    workforce_st_rows = "".join([f"<tr><td>{r['st']}</td><td>{r['total']}</td></tr>" for r in workforce_by_st]) or "<tr><td colspan='2'>Aucune donnée</td></tr>"

    return layout("Tableau de bord", f"""
    <h2>Tableau de bord</h2><div class="help">{help_text}</div>
    <div class="grid">
        <div class="card">Projets<div class="value">{total_projects}</div></div>
        <div class="card">Sous-traitants<div class="value">{total_st}</div></div>
        <div class="card">Produits<div class="value">{total_products}</div></div>
        <div class="card">Stock total<div class="value">{total_stock}</div></div>
    </div><br>
    <div class="grid">
        <div class="card">Entrées totales<div class="value ok">{total_in}</div></div>
        <div class="card">Utilisées totales<div class="value warn">{total_used}</div></div>
        <div class="card">Produits non reconnus<div class="value danger">{unknown}</div></div>
        <div class="card">Saisies sans FDS<div class="value danger">{no_fds}</div></div>
    </div><br>
    <div class="grid">
        <div class="card">Alertes QHSE<div class="value warn">{alerts_count}</div></div>
        <div class="card">Alertes critiques<div class="value danger">{critical_alerts}</div></div>
        <div class="card">Alertes attention<div class="value warn">{warning_alerts}</div></div>
        <div class="card">Version<div class="value ok">2.6</div></div>
    </div><br>
    <div class="grid">
        <div class="card">Effectif aujourd'hui<div class="value ok">{workforce_today}</div></div>
        <div class="card">Effectif semaine<div class="value ok">{workforce_week}</div></div>
        <div class="card">Effectif mois<div class="value ok">{workforce_month}</div></div>
        <div class="card">Effectif année<div class="value ok">{workforce_year}</div></div>
    </div>

    <h3>Analyse rapide</h3>
    <div class="row">
        <div>
            <h3>Top 10 produits stockés</h3>
            <table><tr><th>Produit</th><th>Stock</th></tr>{top_products_rows}</table>
        </div>
        <div>
            <h3>Top 10 produits utilisés</h3>
            <table><tr><th>Produit</th><th>Quantité utilisée</th></tr>{top_used_rows}</table>
        </div>
    </div>

    <div class="row">
        <div>
            <h3>Stock par sous-traitant</h3>
            <table><tr><th>Sous-traitant</th><th>Nb produits</th><th>Stock</th></tr>{st_rows}</table>
            <h3>Effectifs par sous-traitant</h3>
            <table><tr><th>Sous-traitant</th><th>Total période courante</th></tr>{workforce_st_rows}</table>
        </div>
        <div>
            <h3>Répartition par famille</h3>
            <table><tr><th>Famille</th><th>Nb produits</th><th>Stock</th></tr>{family_rows}</table>
        </div>
    </div>

    <h3>Évolution mensuelle</h3>
    <table><tr><th>Mois</th><th>Entrées</th><th>Utilisées</th><th>Stock</th></tr>{month_rows}</table>
    """)

@app.route("/projects", methods=["GET","POST"])
def projects():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        if is_project_admin():
            flash("Un Admin Projet peut modifier son projet depuis la liste, mais ne peut pas créer un nouveau projet.")
            conn.close(); return redirect(url_for("projects"))
        name = request.form["name"].strip()
        conn.execute("""INSERT OR IGNORE INTO projects(name,location,country,client,owner,main_contractor,project_manager,qhse_manager,coordinator,start_date,end_date,status,description)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (name, request.form.get("location",""), request.form.get("country",""), request.form.get("client",""),
                      request.form.get("owner",""), request.form.get("main_contractor",""), request.form.get("project_manager",""),
                      request.form.get("qhse_manager",""), request.form.get("coordinator",""), request.form.get("start_date",""),
                      request.form.get("end_date",""), request.form.get("status","Actif"), request.form.get("description","")))
        conn.commit()
        log_action("Création projet", name)
        flash("Projet enregistré.")
    if is_super_admin():
        rows = conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
    else:
        rows = conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall()
    conn.close()
    trs = ""
    for r in rows:
        edit_btn = f"<a class='btn mini' href='{url_for('edit_project', project_id=r['id'])}'>Modifier</a>" if can_manage_project(r) else ""
        del_btn = f"<form method='post' action='{url_for('delete_project', project_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer ce projet ? Cette action est définitive.&quot;)'><button class='mini btn-danger'>Supprimer</button></form>" if can_delete_project(r) else ""
        trs += f"<tr><td>{r['name']}</td><td>{r['client'] or ''}</td><td>{r['main_contractor'] or ''}</td><td>{r['qhse_manager'] or ''}</td><td>{r['location'] or ''}</td><td>{r['country'] or ''}</td><td>{r['status']}</td><td class='actions'>{edit_btn} {del_btn}</td></tr>"
    add_form = "" if is_project_admin() else f"""<div class="box"><form method="post">
    <div class="row3"><div><label>Nom du projet</label><input name="name" required placeholder="Ex : CHU Kara"></div><div><label>Ville</label><input name="location" placeholder="Ex : Kara"></div><div><label>Pays</label><input name="country" value="Togo" placeholder="Ex : Togo"></div></div>
    <div class="row3"><div><label>Client</label><input name="client" placeholder="Nom du client"></div><div><label>Maître d'ouvrage</label><input name="owner" placeholder="Institution / propriétaire"></div><div><label>Entreprise principale</label><input name="main_contractor" value="Ellipse Projects" placeholder="Entreprise principale"></div></div>
    <div class="row3"><div><label>Chef de projet</label><input name="project_manager" placeholder="Nom du chef de projet"></div><div><label>Responsable QHSE</label><input name="qhse_manager" placeholder="Nom du responsable QHSE"></div><div><label>Coordinateur</label><input name="coordinator" placeholder="Nom du coordinateur"></div></div>
    <div class="row3"><div><label>Date début</label><input type="date" name="start_date"></div><div><label>Date fin</label><input type="date" name="end_date"></div><div><label>Statut</label><select name="status"><option>Actif</option><option>Clôturé</option><option>Inactif</option></select></div></div>
    <label>Description</label><textarea name="description" placeholder="Résumé du projet, lots concernés, observations importantes"></textarea><button>Ajouter</button></form></div>"""
    return layout("Projets", f"""<h2>Projets</h2><div class="help">Chaque ligne peut être modifiée ou supprimée selon ton niveau de droit. Le Super Admin gère tous les projets ; l'Admin Projet ne voit que son projet.</div>{add_form}
    <table><tr><th>Projet</th><th>Client</th><th>Entreprise principale</th><th>Resp. QHSE</th><th>Ville</th><th>Pays</th><th>Statut</th><th>Actions</th></tr>{trs}</table>""")

@app.route("/projects/edit/<int:project_id>", methods=["GET","POST"])
def edit_project(project_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db(); r = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    if not r: conn.close(); flash("Projet introuvable."); return redirect(url_for("projects"))
    if not can_manage_project(r): conn.close(); flash("Accès refusé."); return redirect(url_for("projects"))
    if request.method == "POST":
        conn.execute("""UPDATE projects SET name=?,location=?,country=?,client=?,owner=?,main_contractor=?,project_manager=?,qhse_manager=?,coordinator=?,start_date=?,end_date=?,status=?,description=? WHERE id=?""",
                     (request.form["name"].strip(), request.form.get("location",""), request.form.get("country",""), request.form.get("client",""), request.form.get("owner",""), request.form.get("main_contractor",""), request.form.get("project_manager",""), request.form.get("qhse_manager",""), request.form.get("coordinator",""), request.form.get("start_date",""), request.form.get("end_date",""), request.form.get("status","Actif"), request.form.get("description",""), project_id))
        conn.commit(); conn.close(); log_action("Modification projet", request.form["name"]); flash("Projet modifié."); return redirect(url_for("projects"))
    status = status_options(r['status'])
    body=f"""<h2>Modifier le projet</h2><div class="box"><form method="post">
    <div class="row3"><div><label>Nom du projet</label><input name="name" value="{escape(r['name'])}" required placeholder="Ex : CHU Kara"></div><div><label>Ville</label><input name="location" value="{escape(r['location'] or '')}" placeholder="Ex : Kara"></div><div><label>Pays</label><input name="country" value="{escape(r['country'] or '')}" placeholder="Ex : Togo"></div></div>
    <div class="row3"><div><label>Client</label><input name="client" value="{escape(r['client'] or '')}" placeholder="Nom du client"></div><div><label>Maître d'ouvrage</label><input name="owner" value="{escape(r['owner'] or '')}" placeholder="Institution / propriétaire"></div><div><label>Entreprise principale</label><input name="main_contractor" value="{escape(r['main_contractor'] or '')}" placeholder="Entreprise principale"></div></div>
    <div class="row3"><div><label>Chef de projet</label><input name="project_manager" value="{escape(r['project_manager'] or '')}" placeholder="Nom du chef de projet"></div><div><label>Responsable QHSE</label><input name="qhse_manager" value="{escape(r['qhse_manager'] or '')}" placeholder="Nom du responsable QHSE"></div><div><label>Coordinateur</label><input name="coordinator" value="{escape(r['coordinator'] or '')}" placeholder="Nom du coordinateur"></div></div>
    <div class="row3"><div><label>Date début</label><input type="date" name="start_date" value="{r['start_date'] or ''}"></div><div><label>Date fin</label><input type="date" name="end_date" value="{r['end_date'] or ''}"></div><div><label>Statut</label><select name="status">{status}</select></div></div>
    <label>Description</label><textarea name="description" placeholder="Résumé du projet, lots concernés, observations importantes">{escape(r['description'] or '')}</textarea>
    <button>Enregistrer</button> <a class="btn" href="{url_for('projects')}">Annuler</a></form></div>"""
    conn.close(); return layout("Modifier projet", body)

@app.route("/projects/delete/<int:project_id>", methods=["POST"])
def delete_project(project_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db(); r = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    if not r or not can_delete_project(r): conn.close(); flash("Suppression non autorisée."); return redirect(url_for("projects"))
    if conn.execute("SELECT COUNT(*) c FROM entries WHERE project_id=?", (project_id,)).fetchone()["c"] > 0:
        conn.close(); flash("Impossible de supprimer : ce projet contient déjà des saisies. Passe-le plutôt en Inactif/Clôturé."); return redirect(url_for("projects"))
    conn.execute("DELETE FROM projects WHERE id=?", (project_id,)); conn.commit(); conn.close(); log_action("Suppression projet", r['name']); flash("Projet supprimé."); return redirect(url_for("projects"))

@app.route("/users", methods=["GET","POST"])
def users():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        r = request.form["role"]
        project_id = request.form.get("project_id") or None
        subcontractor_id = request.form.get("subcontractor_id") or None
        if not username or not password:
            flash("Nom d'utilisateur et mot de passe obligatoires."); conn.close(); return redirect(url_for("users"))
        if is_project_admin():
            if r not in ["subcontractor","reader"]:
                flash("Un Admin Projet ne peut créer que des comptes Sous-traitant ou Lecteur."); conn.close(); return redirect(url_for("users"))
            project_id = current_project_id()
        if r == "super_admin":
            project_id = None; subcontractor_id = None
        if r in ["project_admin","reader"]:
            subcontractor_id = None
        if r == "project_admin" and not project_id:
            flash("Choisis un projet pour l'Admin Projet."); conn.close(); return redirect(url_for("users"))
        if r == "subcontractor" and not subcontractor_id:
            flash("Choisis un sous-traitant associé."); conn.close(); return redirect(url_for("users"))
        try:
            conn.execute("INSERT INTO users(username,password_hash,role,subcontractor_id,project_id,status) VALUES(?,?,?,?,?,?)",
                         (username, generate_password_hash(password), r, subcontractor_id, project_id, request.form.get("status","Actif")))
            conn.commit(); log_action("Création utilisateur", username); flash("Utilisateur créé.")
        except sqlite3.IntegrityError:
            flash("Ce nom d'utilisateur existe déjà.")
    page, per_page, offset = pagination_data()
    if is_super_admin():
        projects_rows = conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
        total_users = conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
        users_rows = conn.execute("""SELECT u.*, s.name st_name, p.name project_name FROM users u
                                     LEFT JOIN subcontractors s ON s.id=u.subcontractor_id
                                     LEFT JOIN projects p ON p.id=u.project_id ORDER BY u.username LIMIT ? OFFSET ?""", (per_page, offset)).fetchall()
    else:
        projects_rows = conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall()
        total_users = conn.execute("SELECT COUNT(*) c FROM users WHERE project_id=?", (current_project_id(),)).fetchone()["c"]
        users_rows = conn.execute("""SELECT u.*, s.name st_name, p.name project_name FROM users u
                                     LEFT JOIN subcontractors s ON s.id=u.subcontractor_id
                                     LEFT JOIN projects p ON p.id=u.project_id
                                     WHERE u.project_id=? ORDER BY u.username LIMIT ? OFFSET ?""", (current_project_id(), per_page, offset)).fetchall()
    sts = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall(); conn.close()
    role_options = "<option value='subcontractor'>Sous-traitant</option><option value='reader'>Lecteur</option>"
    if is_super_admin():
        role_options = "<option value='subcontractor'>Sous-traitant</option><option value='project_admin'>Admin Projet</option><option value='reader'>Lecteur</option><option value='super_admin'>Super Admin</option>"
    project_opts = "".join([f"<option value='{p['id']}'>{p['name']}</option>" for p in projects_rows])
    st_opts = "".join([f"<option value='{st['id']}'>{st['name']}</option>" for st in sts])
    trs = ""
    for u in users_rows:
        edit_btn = f"<a class='btn mini' href='{url_for('edit_user', user_id=u['id'])}'>Modifier</a>" if can_manage_user(u) else ""
        del_btn = f"<form method='post' action='{url_for('delete_user', user_id=u['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer cet utilisateur ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>" if can_delete_user(u) else ""
        toggle_label = "Réactiver" if u["status"] == "Inactif" else "Suspendre"
        toggle_class = "" if u["status"] == "Inactif" else "btn-warn"
        toggle_btn = f"<form method='post' action='{url_for('toggle_user_status', user_id=u['id'])}' style='display:inline'><button class='mini {toggle_class}'>{toggle_label}</button></form>" if can_delete_user(u) else ""
        trs += f"<tr><td>{u['username']}</td><td>{u['role']}</td><td>{u['project_name'] or 'Tous / Aucun'}</td><td>{u['st_name'] or '-'}</td><td>{u['status']}</td><td class='actions'>{edit_btn} {toggle_btn} {del_btn}</td></tr>"
    pagination = pagination_html(page, per_page, total_users)
    return layout("Utilisateurs", f"""<h2>Utilisateurs et droits</h2>
    <div class="help"><strong>Correction intégrée :</strong> la création d'utilisateur vérifie maintenant le rôle, le projet et le sous-traitant. Les boutons Modifier/Supprimer apparaissent selon les droits.</div>
    <div class="box"><form method="post">
    <div class="row"><div><label>Nom d'utilisateur</label><input name="username" required placeholder="Ex : sicone_user ou admin_chu_kara"></div><div><label>Mot de passe provisoire</label><input name="password" type="password" required placeholder="Mot de passe temporaire à communiquer à l'utilisateur"></div></div>
    <div class="row"><div><label>Rôle</label><select name="role">{role_options}</select></div><div><label>Projet autorisé</label><select name="project_id"><option value="">Tous / Aucun</option>{project_opts}</select></div></div>
    <div class="row"><div><label>Sous-traitant associé</label><select name="subcontractor_id"><option value="">-- aucun --</option>{st_opts}</select></div><div><label>Statut</label><select name="status"><option>Actif</option><option>Inactif</option></select></div></div>
    <button>Créer l'utilisateur</button></form></div>
    <table><tr><th>Utilisateur</th><th>Rôle</th><th>Projet</th><th>Sous-traitant associé</th><th>Statut</th><th>Actions</th></tr>{trs}</table>{pagination}""")

@app.route("/users/toggle-status/<int:user_id>", methods=["POST"])
def toggle_user_status(user_id):
    if not logged() or not can_access_admin_features():
        return redirect(url_for("dashboard"))
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user or not can_manage_user(user) or str(user_id) == str(session.get("user_id")):
        conn.close(); flash("Suspension non autorisée."); return redirect(url_for("users"))
    new_status = "Actif" if user["status"] == "Inactif" else "Inactif"
    conn.execute("UPDATE users SET status=? WHERE id=?", (new_status, user_id))
    conn.commit(); conn.close()
    log_action("Changement statut utilisateur", f"{user['username']} → {new_status}")
    flash(f"Utilisateur {user['username']} : statut {new_status}.")
    return redirect(url_for("users"))


@app.route("/users/edit/<int:user_id>", methods=["GET","POST"])
def edit_user(user_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db(); u = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not u or not can_manage_user(u): conn.close(); flash("Utilisateur introuvable ou accès refusé."); return redirect(url_for("users"))
    if request.method == "POST":
        r = request.form["role"]; project_id = request.form.get("project_id") or None; subcontractor_id = request.form.get("subcontractor_id") or None
        if is_project_admin():
            if r not in ["subcontractor","reader"]: flash("Rôle non autorisé."); conn.close(); return redirect(url_for("users"))
            project_id = current_project_id()
        if r == "super_admin": project_id = None; subcontractor_id = None
        if r in ["project_admin","reader"]: subcontractor_id = None
        if r == "project_admin" and not project_id: flash("Projet obligatoire pour Admin Projet."); conn.close(); return redirect(url_for("users"))
        if r == "subcontractor" and not subcontractor_id: flash("Sous-traitant obligatoire."); conn.close(); return redirect(url_for("users"))
        if request.form.get("password","").strip():
            conn.execute("UPDATE users SET username=?, password_hash=?, role=?, subcontractor_id=?, project_id=?, status=? WHERE id=?", (request.form["username"].strip(), generate_password_hash(request.form["password"].strip()), r, subcontractor_id, project_id, request.form.get("status","Actif"), user_id))
        else:
            conn.execute("UPDATE users SET username=?, role=?, subcontractor_id=?, project_id=?, status=? WHERE id=?", (request.form["username"].strip(), r, subcontractor_id, project_id, request.form.get("status","Actif"), user_id))
        conn.commit(); conn.close(); log_action("Modification utilisateur", request.form["username"]); flash("Utilisateur modifié."); return redirect(url_for("users"))
    projects_rows = conn.execute("SELECT * FROM projects ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall()
    sts = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall(); conn.close()
    roles = ["subcontractor","reader"] if is_project_admin() else ["subcontractor","project_admin","reader","super_admin"]
    role_options = "".join([f"<option value='{r}' {'selected' if u['role']==r else ''}>{r}</option>" for r in roles])
    project_opts = "".join([f"<option value='{p['id']}' {'selected' if str(u['project_id'])==str(p['id']) else ''}>{p['name']}</option>" for p in projects_rows])
    st_opts = "".join([f"<option value='{st['id']}' {'selected' if str(u['subcontractor_id'])==str(st['id']) else ''}>{st['name']}</option>" for st in sts])
    status = "".join([f"<option {'selected' if u['status']==v else ''}>{v}</option>" for v in ["Actif","Inactif"]])
    return layout("Modifier utilisateur", f"""<h2>Modifier utilisateur</h2><div class="box"><form method="post">
    <div class="row"><div><label>Nom d'utilisateur</label><input name="username" value="{escape(u['username'])}" required placeholder="Identifiant de connexion"></div><div><label>Nouveau mot de passe</label><input name="password" type="password" placeholder="Laisser vide pour conserver l'ancien"></div></div>
    <div class="row"><div><label>Rôle</label><select name="role">{role_options}</select></div><div><label>Projet autorisé</label><select name="project_id"><option value="">Tous / Aucun</option>{project_opts}</select></div></div>
    <div class="row"><div><label>Sous-traitant associé</label><select name="subcontractor_id"><option value="">-- aucun --</option>{st_opts}</select></div><div><label>Statut</label><select name="status">{status}</select></div></div>
    <button>Enregistrer</button> <a class="btn" href="{url_for('users')}">Annuler</a></form></div>""")

@app.route("/users/delete/<int:user_id>", methods=["POST"])
def delete_user(user_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db(); u = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not u or not can_delete_user(u): conn.close(); flash("Suppression non autorisée."); return redirect(url_for("users"))
    conn.execute("DELETE FROM users WHERE id=?", (user_id,)); conn.commit(); conn.close(); log_action("Suppression utilisateur", u['username']); flash("Utilisateur supprimé."); return redirect(url_for("users"))

@app.route("/subcontractors", methods=["GET","POST"])
def subcontractors():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        name = request.form["name"].strip().upper()
        conn.execute("""INSERT OR IGNORE INTO subcontractors(name,contact,phone,email,address,lot,work_zone,employees_count,start_date,end_date,hse_plan,insurance,ppsps,status,observations)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (name, request.form.get("contact",""), request.form.get("phone",""), request.form.get("email",""),
                      request.form.get("address",""), request.form.get("lot",""), request.form.get("work_zone",""), safe_int(request.form.get("employees_count")),
                      request.form.get("start_date",""), request.form.get("end_date",""), request.form.get("hse_plan","NON"), request.form.get("insurance","NON"),
                      request.form.get("ppsps","NON"), request.form.get("status","Actif"), request.form.get("observations","")))
        conn.commit(); log_action("Création sous-traitant", name); flash("Sous-traitant enregistré.")
    rows = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall(); conn.close()
    trs = ""
    for r in rows:
        edit_btn = f"<a class='btn mini' href='{url_for('edit_subcontractor', subcontractor_id=r['id'])}'>Modifier</a>" if can_manage_subcontractor(r) else ""
        del_btn = f"<form method='post' action='{url_for('delete_subcontractor', subcontractor_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer ce sous-traitant ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>" if can_delete_subcontractor(r) else ""
        trs += f"<tr><td>{r['name']}</td><td>{r['contact'] or ''}</td><td>{r['phone'] or ''}</td><td>{r['email'] or ''}</td><td>{r['lot'] or ''}</td><td>{r['work_zone'] or ''}</td><td>{r['employees_count'] or 0}</td><td>{r['hse_plan']}</td><td>{r['insurance']}</td><td>{r['ppsps']}</td><td>{r['status']}</td><td class='actions'>{edit_btn} {del_btn}</td></tr>"
    return layout("Sous-traitants", f"""<h2>Sous-traitants</h2>
    <div class="help"><strong>PPSPS</strong> : Plan Particulier de Sécurité et de Protection de la Santé. <strong>Assurance</strong> : preuve de couverture de l'entreprise. <strong>Plan HSE</strong> : organisation santé, sécurité et environnement du sous-traitant.</div>
    <div class="box"><form method="post">
    <div class="row3"><div><label>Nom</label><input name="name" required placeholder="Ex : SICONE"></div><div><label>Contact</label><input name="contact" placeholder="Nom du responsable / point focal"></div><div><label>Téléphone</label><input name="phone" placeholder="Ex : +228 xx xx xx xx"></div></div>
    <div class="row3"><div><label>Email</label><input name="email" type="email" placeholder="exemple@societe.com"></div><div><label>Lot</label><input name="lot" placeholder="Ex : électricité, plomberie, peinture"></div><div><label>Zone d'intervention</label><input name="work_zone" placeholder="Ex : bâtiment A, magasin, toiture"></div></div>
    <label>Adresse</label><input name="address" placeholder="Adresse complète du sous-traitant">
    <div class="row3"><div><label>Nombre employés</label><input type="number" name="employees_count" value="0" placeholder="Effectif présent sur site"></div><div><label>Date arrivée</label><input type="date" name="start_date"></div><div><label>Date départ</label><input type="date" name="end_date"></div></div>
    <div class="row3"><div><label>Plan HSE</label><select name="hse_plan"><option>OUI</option><option selected>NON</option></select><div class="muted">Document décrivant l'organisation HSE.</div></div><div><label>Assurance</label><select name="insurance"><option>OUI</option><option selected>NON</option></select><div class="muted">Attestation d'assurance valide.</div></div><div><label>PPSPS</label><select name="ppsps"><option>OUI</option><option selected>NON</option></select><div class="muted">Plan sécurité spécifique aux travaux.</div></div></div>
    <label>Statut</label><select name="status"><option>Actif</option><option>Inactif</option></select><label>Observations</label><textarea name="observations" placeholder="Documents manquants, remarques HSE, restrictions d'accès..."></textarea>
    <button>Ajouter</button></form></div>
    <table><tr><th>Nom</th><th>Contact</th><th>Téléphone</th><th>Email</th><th>Lot</th><th>Zone</th><th>Employés</th><th>Plan HSE</th><th>Assurance</th><th>PPSPS</th><th>Statut</th><th>Actions</th></tr>{trs}</table>""")

@app.route("/subcontractors/edit/<int:subcontractor_id>", methods=["GET","POST"])
def edit_subcontractor(subcontractor_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM subcontractors WHERE id=?", (subcontractor_id,)).fetchone()
    if not r or not can_manage_subcontractor(r): conn.close(); flash("Sous-traitant introuvable ou accès refusé."); return redirect(url_for("subcontractors"))
    if request.method=="POST":
        name=request.form["name"].strip().upper()
        conn.execute("""UPDATE subcontractors SET name=?,contact=?,phone=?,email=?,address=?,lot=?,work_zone=?,employees_count=?,start_date=?,end_date=?,hse_plan=?,insurance=?,ppsps=?,status=?,observations=? WHERE id=?""", (name,request.form.get("contact",""),request.form.get("phone",""),request.form.get("email",""),request.form.get("address",""),request.form.get("lot",""),request.form.get("work_zone",""),safe_int(request.form.get("employees_count")),request.form.get("start_date",""),request.form.get("end_date",""),request.form.get("hse_plan","NON"),request.form.get("insurance","NON"),request.form.get("ppsps","NON"),request.form.get("status","Actif"),request.form.get("observations",""),subcontractor_id))
        conn.commit(); conn.close(); log_action("Modification sous-traitant", name); flash("Sous-traitant modifié."); return redirect(url_for("subcontractors"))
    body=f"""<h2>Modifier sous-traitant</h2><div class="box"><form method="post">
    <div class="row3"><div><label>Nom</label><input name="name" value="{escape(r['name'])}" required placeholder="Ex : SICONE"></div><div><label>Contact</label><input name="contact" value="{escape(r['contact'] or '')}" placeholder="Nom du responsable / point focal"></div><div><label>Téléphone</label><input name="phone" value="{escape(r['phone'] or '')}" placeholder="Ex : +228 xx xx xx xx"></div></div>
    <div class="row3"><div><label>Email</label><input name="email" type="email" value="{escape(r['email'] or '')}" placeholder="exemple@societe.com"></div><div><label>Lot</label><input name="lot" value="{escape(r['lot'] or '')}" placeholder="Ex : électricité, plomberie, peinture"></div><div><label>Zone d'intervention</label><input name="work_zone" value="{escape(r['work_zone'] or '')}" placeholder="Ex : bâtiment A, magasin, toiture"></div></div>
    <label>Adresse</label><input name="address" value="{escape(r['address'] or '')}" placeholder="Adresse complète du sous-traitant">
    <div class="row3"><div><label>Nombre employés</label><input type="number" name="employees_count" value="{r['employees_count'] or 0}" placeholder="Effectif présent sur site"></div><div><label>Date arrivée</label><input type="date" name="start_date" value="{r['start_date'] or ''}"></div><div><label>Date départ</label><input type="date" name="end_date" value="{r['end_date'] or ''}"></div></div>
    <div class="row3"><div><label>Plan HSE</label><select name="hse_plan">{yes_no_options(r['hse_plan'])}</select></div><div><label>Assurance</label><select name="insurance">{yes_no_options(r['insurance'])}</select></div><div><label>PPSPS</label><select name="ppsps">{yes_no_options(r['ppsps'])}</select></div></div>
    <label>Statut</label><select name="status">{status_options(r['status'])}</select><label>Observations</label><textarea name="observations" placeholder="Documents manquants, remarques HSE, restrictions d'accès...">{escape(r['observations'] or '')}</textarea>
    <button>Enregistrer</button> <a class="btn" href="{url_for('subcontractors')}">Annuler</a></form></div>"""
    conn.close(); return layout("Modifier sous-traitant", body)

@app.route("/subcontractors/delete/<int:subcontractor_id>", methods=["POST"])
def delete_subcontractor(subcontractor_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM subcontractors WHERE id=?", (subcontractor_id,)).fetchone()
    if not r or not can_delete_subcontractor(r): conn.close(); flash("Suppression non autorisée."); return redirect(url_for("subcontractors"))
    if conn.execute("SELECT COUNT(*) c FROM entries WHERE subcontractor_id=?", (subcontractor_id,)).fetchone()["c"]>0:
        conn.close(); flash("Impossible de supprimer : ce sous-traitant contient déjà des saisies. Passe-le plutôt en Inactif."); return redirect(url_for("subcontractors"))
    conn.execute("DELETE FROM subcontractors WHERE id=?", (subcontractor_id,)); conn.commit(); conn.close(); log_action("Suppression sous-traitant", r['name']); flash("Sous-traitant supprimé."); return redirect(url_for("subcontractors"))

def product_form_html(action_label="Ajouter", r=None):
    r = r or {}
    def v(k, default=""):
        return escape(str(r[k] if k in r.keys() and r[k] is not None else default)) if hasattr(r, 'keys') else default
    fds_val = r['fds'] if hasattr(r,'keys') and 'fds' in r.keys() else 'NON'
    pictos = ["SGH01 Explosif","SGH02 Inflammable","SGH03 Comburant","SGH04 Gaz sous pression","SGH05 Corrosif","SGH06 Toxicité aiguë","SGH07 Irritant / nocif","SGH08 Danger grave pour la santé","SGH09 Danger environnement","SGH02 + SGH07","SGH05 + SGH07"]
    current_picto = r['pictogrammes'] if hasattr(r,'keys') and 'pictogrammes' in r.keys() else ''
    picto_opts = "<option value=''>-- Sélectionner --</option>" + "".join([f"<option {'selected' if current_picto==x else ''}>{x}</option>" for x in pictos])
    return f"""<div class="box"><form method="post">
    <h3>Identification</h3>
    <div class="row3"><div><label>Produit racine</label><input name="root_name" value="{v('root_name')}" required placeholder="Ex : SOPRALENE"></div><div><label>Nom commercial</label><input name="commercial_name" value="{v('commercial_name')}" placeholder="Nom inscrit sur le contenant"></div><div><label>Fabricant</label><input name="manufacturer" value="{v('manufacturer')}" placeholder="Fabricant / fournisseur"></div></div>
    <div class="row3"><div><label>Référence</label><input name="reference" value="{v('reference')}" placeholder="Référence produit, code fournisseur ou CAS"></div><div><label>Famille</label><input name="family" value="{v('family')}" placeholder="Peinture, colle, solvant, étanchéité..."></div><div><label>État physique</label><input name="physical_state" value="{v('physical_state')}" placeholder="Liquide, solide, pâte, poudre, gaz..."></div></div>
    <label>Conditionnement</label><input name="conditionnement" value="{v('conditionnement')}" placeholder="Ex : seau 20 kg, bidon 5 L, sac 25 kg"><label>Utilisation</label><input name="utilisation" value="{v('utilisation')}" placeholder="Ex : peinture de finition, étanchéité toiture...">
    <h3>FDS et sécurité</h3>
    <div class="row3"><div><label>FDS</label><select name="fds">{yes_no_options(fds_val)}</select></div><div><label>Date FDS</label><input type="date" name="fds_date" value="{v('fds_date')}" placeholder="Date d'émission de la FDS"></div><div><label>Expiration FDS</label><input type="date" name="fds_expiry" value="{v('fds_expiry')}" placeholder="Date de fin de validité ou prochaine révision"></div></div>
    <div class="row3"><div><label>Version FDS</label><input name="fds_version" value="{v('fds_version')}" placeholder="Ex : Version 1.0 / Révision 2026"></div><div><label>Pictogrammes</label><select name="pictogrammes">{picto_opts}</select></div><div><label>Classe de danger</label><input name="danger_class" value="{v('danger_class')}" placeholder="Ex : inflammable, corrosif, toxique..."></div></div>
    <label>Mentions H</label><input name="h_statements" value="{v('h_statements')}" placeholder="Ex : H225 – Liquide et vapeurs très inflammables">
    <label>Mentions P</label><input name="p_statements" value="{v('p_statements')}" placeholder="Ex : P280 – Porter des gants de protection">
    <label>Incompatibilités</label><input name="incompatibilites" value="{v('incompatibilites')}" placeholder="Produits à éviter : acides, bases, oxydants...">
    <label>EPI requis</label><input name="epi" value="{v('epi')}" placeholder="Gants, lunettes, masque, combinaison...">
    <label>Premiers secours</label><textarea name="first_aid" placeholder="Mesures à appliquer en cas d'exposition, inhalation, contact peau/yeux...">{v('first_aid')}</textarea>
    <label>Mesures incendie</label><textarea name="fire_measures" placeholder="Moyens d'extinction appropriés/interdits, risques particuliers...">{v('fire_measures')}</textarea>
    <label>Déversement accidentel</label><textarea name="spill_measures" placeholder="Procédure de confinement, récupération, nettoyage et élimination...">{v('spill_measures')}</textarea>
    <label>Règles de stockage</label><textarea name="storage_rules" placeholder="Conditions de stockage, ventilation, séparation des incompatibles...">{v('storage_rules')}</textarea>
    <h3>Seuils de stock</h3><div class="row"><div><label>Stock minimum</label><input type="number" step="0.01" name="stock_min" value="{v('stock_min','0')}" placeholder="Quantité minimale avant alerte"></div><div><label>Stock maximum</label><input type="number" step="0.01" name="stock_max" value="{v('stock_max','0')}" placeholder="Quantité maximale autorisée"></div></div>
    <label>Observations</label><textarea name="observations" placeholder="Remarques complémentaires, restrictions, suivi audit...">{v('observations')}</textarea><button>{action_label}</button></form></div>"""

def save_product_from_form(conn, product_id=None):
    data = (request.form["root_name"].strip().upper(), request.form.get("commercial_name",""), request.form.get("manufacturer",""), request.form.get("reference",""),
         request.form.get("family",""), request.form.get("physical_state",""), request.form.get("conditionnement",""), request.form.get("utilisation",""),
         request.form.get("fds","NON"), request.form.get("fds_date",""), request.form.get("fds_version",""), request.form.get("fds_expiry",""),
         request.form.get("pictogrammes",""), request.form.get("danger_class",""), request.form.get("h_statements",""), request.form.get("p_statements",""),
         request.form.get("incompatibilites",""), request.form.get("epi",""), request.form.get("first_aid",""), request.form.get("fire_measures",""),
         request.form.get("spill_measures",""), request.form.get("storage_rules",""), request.form.get("stock_min") or 0, request.form.get("stock_max") or 0,
         request.form.get("observations",""))
    if product_id:
        conn.execute("""UPDATE products SET root_name=?,commercial_name=?,manufacturer=?,reference=?,family=?,physical_state=?,conditionnement=?,utilisation=?,fds=?,fds_date=?,fds_version=?,fds_expiry=?,pictogrammes=?,danger_class=?,h_statements=?,p_statements=?,incompatibilites=?,epi=?,first_aid=?,fire_measures=?,spill_measures=?,storage_rules=?,stock_min=?,stock_max=?,observations=? WHERE id=?""", data + (product_id,))
    else:
        conn.execute("""INSERT OR IGNORE INTO products(root_name,commercial_name,manufacturer,reference,family,physical_state,conditionnement,utilisation,fds,fds_date,fds_version,fds_expiry,pictogrammes,danger_class,h_statements,p_statements,incompatibilites,epi,first_aid,fire_measures,spill_measures,storage_rules,stock_min,stock_max,observations)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", data)

@app.route("/products", methods=["GET","POST"])
def products():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        save_product_from_form(conn)
        conn.commit(); log_action("Création produit", request.form["root_name"]); flash("Produit enregistré.")
    page, per_page, offset = pagination_data()
    total_products = conn.execute("SELECT COUNT(*) c FROM products").fetchone()["c"]
    rows = conn.execute("SELECT * FROM products ORDER BY root_name LIMIT ? OFFSET ?", (per_page, offset)).fetchall(); conn.close()
    trs = ""
    for r in rows:
        edit_btn = f"<a class='btn mini' href='{url_for('edit_product', product_id=r['id'])}'>Modifier</a>"
        del_btn = f"<form method='post' action='{url_for('delete_product', product_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer ce produit ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>"
        trs += f"<tr><td>{r['root_name']}</td><td>{r['family'] or ''}</td><td>{r['conditionnement'] or ''}</td><td>{r['utilisation'] or ''}</td><td>{r['fds']}</td><td>{r['fds_expiry'] or ''}</td><td>{r['pictogrammes'] or ''}</td><td>{r['incompatibilites'] or ''}</td><td>{r['stock_min'] or 0}</td><td>{r['stock_max'] or 0}</td><td class='actions'>{edit_btn} {del_btn}</td></tr>"
    pagination = pagination_html(page, per_page, total_products)
    return layout("Base produits", f"""<h2>Base produits chimiques</h2><div class="help">Tous les champs de saisie contiennent désormais des indications. Les produits peuvent être modifiés ou supprimés selon les droits d'administration.</div>{product_form_html()}
    <table><tr><th>Produit</th><th>Famille</th><th>Conditionnement</th><th>Utilisation</th><th>FDS</th><th>Expiration FDS</th><th>Pictogrammes</th><th>Incompatibilités</th><th>Stock min</th><th>Stock max</th><th>Actions</th></tr>{trs}</table>{pagination}""")

@app.route("/products/edit/<int:product_id>", methods=["GET","POST"])
def edit_product(product_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not r: conn.close(); flash("Produit introuvable."); return redirect(url_for("products"))
    if request.method=="POST":
        save_product_from_form(conn, product_id); conn.commit(); conn.close(); log_action("Modification produit", request.form["root_name"]); flash("Produit modifié."); return redirect(url_for("products"))
    conn.close(); return layout("Modifier produit", f"<h2>Modifier produit chimique</h2>{product_form_html('Enregistrer les modifications', r)}<p><a class='btn' href='{url_for('products')}'>Retour</a></p>")

@app.route("/products/delete/<int:product_id>", methods=["POST"])
def delete_product(product_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not r: conn.close(); flash("Produit introuvable."); return redirect(url_for("products"))
    if conn.execute("SELECT COUNT(*) c FROM entries WHERE product_id=?", (product_id,)).fetchone()["c"]>0:
        conn.close(); flash("Impossible de supprimer : ce produit est utilisé dans des saisies. Modifie-le ou crée un alias."); return redirect(url_for("products"))
    conn.execute("DELETE FROM aliases WHERE product_id=?", (product_id,)); conn.execute("DELETE FROM products WHERE id=?", (product_id,)); conn.commit(); conn.close(); log_action("Suppression produit", r['root_name']); flash("Produit supprimé."); return redirect(url_for("products"))

@app.route("/aliases", methods=["GET","POST"])
def aliases():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        alias_name = normalize(request.form.get("alias_name", ""))
        product_id = request.form.get("product_id")
        if not alias_name or not product_id:
            flash("Alias non enregistré : le nom de l’alias et le produit racine sont obligatoires.")
        else:
            try:
                conn.execute("INSERT INTO aliases(alias_name,product_id) VALUES(?,?)", (alias_name, product_id))
                conn.commit(); log_action("Création alias", alias_name); flash("Alias enregistré.")
            except sqlite3.IntegrityError:
                flash("Cet alias existe déjà. Utilise le bouton Modifier pour le corriger ou le rattacher à un autre produit racine.")
    products = conn.execute("SELECT * FROM products ORDER BY root_name").fetchall()
    aliases_rows = conn.execute("SELECT a.id,a.alias_name,a.product_id,p.root_name FROM aliases a JOIN products p ON p.id=a.product_id ORDER BY a.alias_name").fetchall(); conn.close()
    opts = "".join([f"<option value='{p['id']}'>{p['root_name']}</option>" for p in products])
    trs = ""
    for a in aliases_rows:
        trs += f"""<tr>
            <td>{escape(a['alias_name'])}</td>
            <td>{escape(a['root_name'])}</td>
            <td class='actions'>
                <a class='btn mini' href='{url_for('edit_alias', alias_id=a['id'])}'>Modifier</a>
                <form method='post' action='{url_for('delete_alias', alias_id=a['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer cet alias/doublon ? Cette action retirera uniquement l’alias, pas le produit racine.&quot;)'>
                    <button type='submit' class='mini btn-danger'>Supprimer</button>
                </form>
            </td>
        </tr>"""
    return layout("Alias", f"""<h2>Alias / Doublons</h2>
    <div class="help"><b>Alias / doublon :</b> permet de rattacher une appellation erronée, commerciale ou abrégée à un produit racine. En cas d’erreur de saisie, utiliser <b>Modifier</b>. Si l’alias n’est plus utile, utiliser <b>Supprimer</b>.</div>
    <div class="box"><form method="post"><label>Nom saisi / alias</label><input name="alias_name" required placeholder="Ex : PANTEX VELOUR 20KG, Pantax velour, PANTEX VEL"><label>Produit racine</label><select name="product_id">{opts}</select><button>Créer l'alias</button></form></div>
    <table><tr><th>Alias / doublon saisi</th><th>Produit racine rattaché</th><th>Modifier / Supprimer</th></tr>{trs}</table>""")

@app.route("/aliases/edit/<int:alias_id>", methods=["GET","POST"])
def edit_alias(alias_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); a=conn.execute("SELECT * FROM aliases WHERE id=?", (alias_id,)).fetchone()
    if not a: conn.close(); flash("Alias introuvable."); return redirect(url_for("aliases"))
    if request.method=="POST":
        alias_name = normalize(request.form.get("alias_name", ""))
        product_id = request.form.get("product_id")
        if not alias_name or not product_id:
            flash("Modification impossible : le nom de l’alias et le produit racine sont obligatoires.")
        else:
            try:
                conn.execute("UPDATE aliases SET alias_name=?, product_id=? WHERE id=?", (alias_name, product_id, alias_id))
                conn.commit(); conn.close(); log_action("Modification alias", alias_name); flash("Alias modifié."); return redirect(url_for("aliases"))
            except sqlite3.IntegrityError:
                flash("Modification impossible : un autre alias porte déjà ce nom.")
    products=conn.execute("SELECT * FROM products ORDER BY root_name").fetchall(); conn.close()
    opts="".join([f"<option value='{p['id']}' {'selected' if str(a['product_id'])==str(p['id']) else ''}>{p['root_name']}</option>" for p in products])
    return layout("Modifier alias", f"""<h2>Modifier alias / doublon</h2><div class="help">Corrige ici une erreur de saisie ou rattache l’alias à un autre produit racine.</div><div class="box"><form method="post"><label>Nom saisi / alias</label><input name="alias_name" value="{escape(a['alias_name'])}" required placeholder="Ex : PANTEX VELOUR 20KG"><label>Produit racine</label><select name="product_id">{opts}</select><button>Enregistrer</button> <a class="btn" href="{url_for('aliases')}">Annuler</a></form></div>""")

@app.route("/aliases/delete/<int:alias_id>", methods=["POST"])
def delete_alias(alias_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); a=conn.execute("SELECT * FROM aliases WHERE id=?", (alias_id,)).fetchone()
    if not a: conn.close(); flash("Alias introuvable."); return redirect(url_for("aliases"))
    conn.execute("DELETE FROM aliases WHERE id=?", (alias_id,)); conn.commit(); conn.close(); log_action("Suppression alias", a['alias_name']); flash("Alias supprimé. Le produit racine reste conservé."); return redirect(url_for("aliases"))

@app.route("/entries", methods=["GET","POST"])
def entries():
    if not logged() or is_reader(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        qin = float(request.form.get("qty_in") or 0)
        qused = float(request.form.get("qty_used") or 0)
        qstock = float(request.form.get("qty_stock") or (qin - qused))
        selected_product_id = request.form.get("product_id")
        declared_new = request.form.get("declared_product_new","").strip()

        if selected_product_id:
            product_id = int(selected_product_id)
            p = conn.execute("SELECT root_name,fds,utilisation FROM products WHERE id=?", (product_id,)).fetchone()
            declared_product = p["root_name"] if p else declared_new
            default_fds = p["fds"] if p else "NON"
        else:
            declared_product = declared_new
            product_id = find_product_id(declared_product)
            p = conn.execute("SELECT fds,utilisation FROM products WHERE id=?", (product_id,)).fetchone() if product_id else None
            default_fds = p["fds"] if p else "NON"
        fds_available = request.form.get("fds_available") or default_fds
        utilisation = request.form.get("utilisation", "").strip() or (p["utilisation"] if p and "utilisation" in p.keys() else "")
        if product_id and utilisation:
            conn.execute("UPDATE products SET utilisation=? WHERE id=?", (utilisation, product_id))

        if can_access_admin_features():
            subcontractor_id = request.form["subcontractor_id"]
            project_id = request.form["project_id"]
        else:
            subcontractor_id = current_subcontractor_id()
            project_id = current_project_id() or 1
        if is_project_admin() and str(project_id) != str(current_project_id()):
            flash("Tu ne peux saisir que pour ton projet autorisé.")
            return redirect(url_for("entries"))
        if not declared_product:
            flash("Choisis un produit validé ou saisis un nouveau produit déclaré.")
            return redirect(url_for("entries"))

        conn.execute("""INSERT INTO entries(project_id,entry_date,month,subcontractor_id,declared_product,product_id,utilisation,qty_in,qty_used,qty_stock,unit,storage_location,fds_available,observations)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (project_id, request.form["entry_date"], request.form["month"], subcontractor_id, declared_product, product_id, utilisation,
                      qin, qused, qstock, request.form.get("unit","u"), request.form.get("storage_location",""), fds_available, request.form.get("observations","")))
        conn.commit()
        log_action("Saisie produit", declared_product)
        flash("Saisie enregistrée.")

    products = conn.execute("SELECT * FROM products ORDER BY root_name").fetchall()
    product_opts = "".join([f"<option value='{p['id']}' data-utilisation='{escape(p['utilisation'] or '')}'>{p['root_name']} ({p['family'] or 'Famille non définie'})</option>" for p in products])

    if can_access_admin_features():
        projects = conn.execute("SELECT * FROM projects WHERE status='Actif' ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall()
        project_opts = "".join([f"<option value='{p['id']}'>{p['name']}</option>" for p in projects])
        sts = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall()
        st_opts = "".join([f"<option value='{s['id']}'>{s['name']}</option>" for s in sts])
        admin_fields = f"""<div class="row"><div><label>Projet</label><select name="project_id">{project_opts}</select></div><div><label>Sous-traitant</label><select name="subcontractor_id">{st_opts}</select></div></div>"""
        where, params = where_scope("e")
        page, per_page, offset = pagination_data()
        total_entries = conn.execute(f"SELECT COUNT(*) c FROM entries e {where}", params).fetchone()["c"]
        rows = conn.execute(f"""SELECT e.*,s.name st,p.root_name,COALESCE(e.utilisation,p.utilisation) utilisation,pr.name project_name FROM entries e
                               JOIN subcontractors s ON s.id=e.subcontractor_id
                               LEFT JOIN products p ON p.id=e.product_id
                               LEFT JOIN projects pr ON pr.id=e.project_id
                               {where} ORDER BY e.entry_date DESC,e.id DESC LIMIT ? OFFSET ?""", params + [per_page, offset]).fetchall()
    else:
        admin_fields = "<div class='help'>Tu saisis uniquement pour ton entreprise et ton projet.</div>"
        page, per_page, offset = pagination_data()
        total_entries = conn.execute("SELECT COUNT(*) c FROM entries WHERE subcontractor_id=?", (current_subcontractor_id(),)).fetchone()["c"]
        rows = conn.execute("""SELECT e.*,s.name st,p.root_name,COALESCE(e.utilisation,p.utilisation) utilisation,pr.name project_name FROM entries e
                               JOIN subcontractors s ON s.id=e.subcontractor_id
                               LEFT JOIN products p ON p.id=e.product_id
                               LEFT JOIN projects pr ON pr.id=e.project_id
                               WHERE e.subcontractor_id=?
                               ORDER BY e.entry_date DESC,e.id DESC LIMIT ? OFFSET ?""", (current_subcontractor_id(), per_page, offset)).fetchall()
    conn.close()

    trs = "".join([f"<tr><td>{r['entry_date']}</td><td>{r['month']}</td><td>{r['project_name'] or ''}</td><td>{r['st']}</td><td>{r['declared_product']}</td><td>{r['root_name'] or '<span class=danger>Non reconnu</span>'}</td><td>{r['utilisation'] or ''}</td><td>{r['qty_in']}</td><td>{r['qty_used']}</td><td>{r['qty_stock']}</td><td>{r['unit']}</td><td>{r['fds_available']}</td><td>{r['storage_location'] or ''}</td><td><a class='btn' href='{url_for('edit_entry', entry_id=r['id'])}'>Modifier</a> <form method='post' action='{url_for('delete_entry', entry_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer cette saisie ?&quot;)'><button>Supprimer</button></form></td></tr>" for r in rows])
    pagination = pagination_html(page, per_page, total_entries)
    return layout("Saisie", f"""<h2>Saisie mensuelle</h2><div class="help">Choisis de préférence un produit validé. Le stock est calculé automatiquement si tu laisses le champ stock vide.</div>
    <div class="box"><form method="post">{admin_fields}
    <div class="row"><div><label>Date</label><input type="date" name="entry_date" required></div><div><label>Mois</label><input name="month" placeholder="2026-06" required></div></div>
    <label>Produit validé</label><select name="product_id"><option value="">-- Nouveau produit / non listé --</option>{product_opts}</select>
    <label>Nouveau produit déclaré</label><input name="declared_product_new" placeholder="À remplir seulement si le produit n'est pas dans la liste">
    <label>Utilisation</label><input id="utilisation" name="utilisation" placeholder="Ex : nettoyage, peinture, étanchéité, collage...">
    <div class="row"><div><label>Quantité entrée</label><input id="qty_in" type="number" step="0.01" name="qty_in" value="0" placeholder="Quantité reçue"></div><div><label>Quantité utilisée</label><input id="qty_used" type="number" step="0.01" name="qty_used" value="0" placeholder="Quantité consommée"></div></div>
    <div class="row"><div><label>Stock restant</label><input id="qty_stock" type="number" step="0.01" name="qty_stock" placeholder="Calcul automatique : entrée - utilisée"></div><div><label>Unité</label><select name="unit"><option>u</option><option>kg</option><option>L</option><option>seau</option><option>sac</option><option>rouleau</option><option>bidon</option></select></div></div>
    <label>Lieu de stockage</label><input name="storage_location" placeholder="Ex : magasin peinture, local étanchéité"><label>FDS disponible</label><select name="fds_available"><option>OUI</option><option selected>NON</option></select>
    <label>Observations</label><textarea name="observations"></textarea><button>Enregistrer</button></form></div>
    <table><tr><th>Date</th><th>Mois</th><th>Projet</th><th>Sous-traitant</th><th>Produit déclaré</th><th>Produit racine</th><th>Utilisation</th><th>Entrée</th><th>Utilisée</th><th>Stock</th><th>Unité</th><th>FDS</th><th>Stockage</th><th>Actions</th></tr>{trs}</table>{pagination}
    <script>
    const productSelect=document.querySelector('select[name="product_id"]'); const utilisation=document.getElementById('utilisation');
    if(productSelect&&utilisation){{productSelect.addEventListener('change',()=>{{const o=productSelect.options[productSelect.selectedIndex]; if(o&&o.dataset.utilisation) utilisation.value=o.dataset.utilisation;}});}}
    const qi=document.getElementById('qty_in'); const qu=document.getElementById('qty_used'); const qs=document.getElementById('qty_stock');
    function calcStock(){{ if(!qi||!qu||!qs) return; const a=parseFloat(qi.value)||0; const b=parseFloat(qu.value)||0; qs.value=(a-b).toFixed(2); }}
    if(qi&&qu&&qs){{ qi.addEventListener('input', calcStock); qu.addEventListener('input', calcStock); }}
    </script>""")


@app.route("/entries/edit/<int:entry_id>", methods=["GET","POST"])
def edit_entry(entry_id):
    if not logged() or is_reader():
        return redirect(url_for("dashboard"))
    conn = db()
    entry = conn.execute("SELECT * FROM entries WHERE id=?", (entry_id,)).fetchone()
    if not entry:
        conn.close(); flash("Saisie introuvable."); return redirect(url_for("entries"))
    if not can_manage_entry(entry):
        conn.close(); flash("Accès refusé."); return redirect(url_for("entries"))
    if request.method == "POST":
        qin = safe_float(request.form.get("qty_in"))
        qused = safe_float(request.form.get("qty_used"))
        qstock = safe_float(request.form.get("qty_stock"), qin - qused)
        utilisation = request.form.get("utilisation", "").strip()
        conn.execute("""UPDATE entries SET entry_date=?, month=?, utilisation=?, qty_in=?, qty_used=?, qty_stock=?, unit=?, storage_location=?, fds_available=?, observations=? WHERE id=?""",
                     (request.form["entry_date"], request.form["month"], utilisation, qin, qused, qstock, request.form.get("unit","u"), request.form.get("storage_location",""), request.form.get("fds_available","NON"), request.form.get("observations",""), entry_id))
        if entry["product_id"] and utilisation:
            conn.execute("UPDATE products SET utilisation=? WHERE id=?", (utilisation, entry["product_id"]))
        conn.commit(); conn.close()
        log_action("Modification saisie", f"Saisie ID {entry_id}")
        flash("Saisie modifiée.")
        return redirect(url_for("entries"))
    conn.close()
    unit_options = "".join([f"<option {'selected' if entry['unit']==u else ''}>{u}</option>" for u in ['u','kg','L','seau','sac','rouleau','bidon']])
    fds_options = f"<option {'selected' if entry['fds_available']=='OUI' else ''}>OUI</option><option {'selected' if entry['fds_available']=='NON' else ''}>NON</option>"
    return layout("Modifier une saisie", f"""<h2>Modifier une saisie</h2>
    <div class="box"><form method="post">
    <div class="row"><div><label>Date</label><input type="date" name="entry_date" value="{entry['entry_date']}" required></div>
    <div><label>Mois</label><input name="month" value="{entry['month']}" required placeholder="Ex : 2026-06"></div></div>
    <label>Utilisation</label><input name="utilisation" value="{escape(entry['utilisation'] or '')}" placeholder="Utilisation réelle du produit">
    <div class="row"><div><label>Quantité entrée</label><input id="edit_qty_in" type="number" step="0.01" name="qty_in" value="{entry['qty_in']}" placeholder="Quantité reçue"></div>
    <div><label>Quantité utilisée</label><input id="edit_qty_used" type="number" step="0.01" name="qty_used" value="{entry['qty_used']}" placeholder="Quantité consommée"></div></div>
    <div class="row"><div><label>Stock restant</label><input id="edit_qty_stock" type="number" step="0.01" name="qty_stock" value="{entry['qty_stock']}" placeholder="Calcul automatique"></div>
    <div><label>Unité</label><select name="unit">{unit_options}</select></div></div>
    <label>Lieu de stockage</label><input name="storage_location" value="{entry['storage_location'] or ''}" placeholder="Ex : magasin peinture, local étanchéité">
    <label>FDS disponible</label><select name="fds_available">{fds_options}</select>
    <label>Observations</label><textarea name="observations" placeholder="Commentaires, anomalies, remarques">{entry['observations'] or ''}</textarea>
    <button>Enregistrer les modifications</button> <a class="btn" href="{url_for('entries')}">Annuler</a>
    </form></div>
    <script>
    const qi=document.getElementById('edit_qty_in'); const qu=document.getElementById('edit_qty_used'); const qs=document.getElementById('edit_qty_stock');
    function calcStock(){{ const a=parseFloat(qi.value)||0; const b=parseFloat(qu.value)||0; qs.value=(a-b).toFixed(2); }}
    qi.addEventListener('input', calcStock); qu.addEventListener('input', calcStock);
    </script>""")


@app.route("/entries/delete/<int:entry_id>", methods=["POST"])
def delete_entry(entry_id):
    if not logged() or is_reader():
        return redirect(url_for("dashboard"))
    conn = db()
    entry = conn.execute("SELECT * FROM entries WHERE id=?", (entry_id,)).fetchone()
    if not entry:
        conn.close(); flash("Saisie introuvable."); return redirect(url_for("entries"))
    if not can_manage_entry(entry):
        conn.close(); flash("Accès refusé."); return redirect(url_for("entries"))
    conn.execute("DELETE FROM entries WHERE id=?", (entry_id,)); conn.commit(); conn.close()
    log_action("Suppression saisie", f"Saisie ID {entry_id}")
    flash("Saisie supprimée.")
    return redirect(url_for("entries"))

@app.route("/workforce", methods=["GET","POST"])
def workforce():
    if not logged(): return redirect(url_for("login"))
    conn = db()
    if request.method == "POST":
        work_date = request.form.get("work_date") or date.today().strftime("%Y-%m-%d")
        project_id = request.form.get("project_id") or None
        subcontractor_id = request.form.get("subcontractor_id") or None
        men = safe_int(request.form.get("men"), 0)
        women = safe_int(request.form.get("women"), 0)
        total = men + women
        if is_project_admin(): project_id = current_project_id()
        if is_subcontractor(): subcontractor_id = current_subcontractor_id()
        if is_reader():
            flash("Compte en lecture seule : saisie non autorisée."); conn.close(); return redirect(url_for("workforce"))
        if not subcontractor_id:
            flash("Choisis le sous-traitant."); conn.close(); return redirect(url_for("workforce"))
        if not project_id:
            flash("Choisis le projet/site."); conn.close(); return redirect(url_for("workforce"))
        conn.execute("""INSERT INTO workforce(work_date,project_id,subcontractor_id,men,women,total,observations,created_by,created_at)
                        VALUES(?,?,?,?,?,?,?,?,?)""", (work_date, project_id, subcontractor_id, men, women, total, request.form.get("observations",""), session.get("username",""), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); log_action("Saisie effectif", f"{work_date} - ST {subcontractor_id} - total {total}"); flash("Effectif enregistré.")

    if is_super_admin():
        projects = conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
    else:
        projects = conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall() if current_project_id() else conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
    if is_subcontractor():
        sts = conn.execute("SELECT * FROM subcontractors WHERE id=?", (current_subcontractor_id(),)).fetchall()
    else:
        sts = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall()

    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    clauses=[]; params=[]
    if start_date: clauses.append("w.work_date>=?"); params.append(start_date)
    if end_date: clauses.append("w.work_date<=?"); params.append(end_date)
    scope_where, scope_params = workforce_scope("w")
    if scope_where:
        clauses.append(scope_where.replace(" WHERE ","").strip())
        params += scope_params
    where_sql = " WHERE " + " AND ".join(clauses) if clauses else ""

    page, per_page, offset = pagination_data()
    total_workforce = conn.execute(f"SELECT COUNT(*) c FROM workforce w {where_sql}", params).fetchone()["c"]
    rows = conn.execute(f"""SELECT w.*, p.name project_name, s.name st_name
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id JOIN subcontractors s ON s.id=w.subcontractor_id
        {where_sql} ORDER BY w.work_date DESC, p.name, s.name LIMIT ? OFFSET ?""", params + [per_page, offset]).fetchall()
    summary_day = conn.execute(f"""SELECT w.work_date, COALESCE(p.name,'') project_name, SUM(w.total) total
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id {where_sql}
        GROUP BY w.work_date, project_name ORDER BY w.work_date DESC LIMIT 60""", params).fetchall()
    conn.close()

    project_opts = "".join([f"<option value='{p['id']}'>{p['name']}</option>" for p in projects])
    st_opts = "".join([f"<option value='{st['id']}'>{st['name']}</option>" for st in sts])
    add_form = "" if is_reader() else f"""<div class="box"><form method="post">
        <div class="row3"><div><label>Date</label><input type="date" name="work_date" value="{date.today().strftime('%Y-%m-%d')}" required></div><div><label>Projet / Site</label><select name="project_id">{project_opts}</select></div><div><label>Sous-traitant</label><select name="subcontractor_id">{st_opts}</select></div></div>
        <div class="row3"><div><label>Hommes</label><input type="number" min="0" name="men" value="0"></div><div><label>Femmes</label><input type="number" min="0" name="women" value="0"></div><div><label>Observations</label><input name="observations" placeholder="RAS, activité particulière, remarque HSE..."></div></div>
        <button>Enregistrer l'effectif</button></form></div>"""
    export_btn = f"<a class='btn' href='{url_for('export_workforce_excel')}'>Exporter les effectifs en Excel</a>" if can_access_admin_features() else ""
    trs = ""
    for r in rows:
        actions = ""
        if can_manage_workforce(r) and not is_reader():
            actions = f"<a class='btn mini' href='{url_for('edit_workforce', workforce_id=r['id'])}'>Modifier</a> <form method='post' action='{url_for('delete_workforce', workforce_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer cette saisie d’effectif ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>"
        trs += f"<tr><td>{r['work_date']}</td><td>{r['project_name'] or ''}</td><td>{r['st_name']}</td><td>{r['men']}</td><td>{r['women']}</td><td><strong>{r['total']}</strong></td><td>{r['observations'] or ''}</td><td class='actions'>{actions}</td></tr>"
    trs = trs or "<tr><td colspan='8'>Aucune donnée</td></tr>"
    pagination = pagination_html(page, per_page, total_workforce)
    sum_rows = "".join([f"<tr><td>{r['work_date']}</td><td>{r['project_name']}</td><td><strong>{r['total']}</strong></td></tr>" for r in summary_day]) or "<tr><td colspan='3'>Aucune donnée</td></tr>"
    return layout("Effectifs", f"""<h2>Effectifs journaliers sous-traitants</h2>
    <div class="help">Chaque sous-traitant voit ses propres effectifs. Les admins voient les données selon leurs projets autorisés. Les totaux se compilent automatiquement par jour, semaine, mois et année dans le tableau de bord.</div>
    {export_btn}<br><br>{add_form}
    <div class="box"><form method="get"><div class="row"><div><label>Date début</label><input type="date" name="start_date" value="{start_date}"></div><div><label>Date fin</label><input type="date" name="end_date" value="{end_date}"></div></div><button>Filtrer</button></form></div>
    <h3>Saisies détaillées</h3><table><tr><th>Date</th><th>Projet/Site</th><th>Sous-traitant</th><th>Hommes</th><th>Femmes</th><th>Total</th><th>Observations</th><th>Actions</th></tr>{trs}</table>{pagination}
    <h3>Total journalier par site</h3><table><tr><th>Date</th><th>Site</th><th>Total</th></tr>{sum_rows}</table>""")

@app.route("/workforce/edit/<int:workforce_id>", methods=["GET", "POST"])
def edit_workforce(workforce_id):
    if not logged() or is_reader():
        return redirect(url_for("dashboard"))
    conn = db()
    row = conn.execute("SELECT * FROM workforce WHERE id=?", (workforce_id,)).fetchone()
    if not row or not can_manage_workforce(row):
        conn.close(); flash("Saisie d’effectif introuvable ou accès refusé."); return redirect(url_for("workforce"))
    if request.method == "POST":
        men = safe_int(request.form.get("men"), 0)
        women = safe_int(request.form.get("women"), 0)
        conn.execute("UPDATE workforce SET work_date=?, men=?, women=?, total=?, observations=? WHERE id=?",
                     (request.form.get("work_date"), men, women, men + women, request.form.get("observations", ""), workforce_id))
        conn.commit(); conn.close()
        log_action("Modification effectif", f"Saisie ID {workforce_id}")
        flash("Saisie d’effectif modifiée.")
        return redirect(url_for("workforce"))
    conn.close()
    return layout("Modifier effectif", f"""<h2>Modifier une saisie d’effectif</h2><div class='box'><form method='post'>
    <label>Date</label><input type='date' name='work_date' value='{row['work_date']}' required>
    <div class='row'><div><label>Hommes</label><input type='number' min='0' name='men' value='{row['men']}'></div><div><label>Femmes</label><input type='number' min='0' name='women' value='{row['women']}'></div></div>
    <label>Observations</label><textarea name='observations'>{escape(row['observations'] or '')}</textarea>
    <button>Enregistrer</button> <a class='btn' href='{url_for('workforce')}'>Annuler</a></form></div>""")


@app.route("/workforce/delete/<int:workforce_id>", methods=["POST"])
def delete_workforce(workforce_id):
    if not logged() or is_reader():
        return redirect(url_for("dashboard"))
    conn = db()
    row = conn.execute("SELECT * FROM workforce WHERE id=?", (workforce_id,)).fetchone()
    if not row or not can_manage_workforce(row):
        conn.close(); flash("Suppression non autorisée."); return redirect(url_for("workforce"))
    conn.execute("DELETE FROM workforce WHERE id=?", (workforce_id,))
    conn.commit(); conn.close()
    log_action("Suppression effectif", f"Saisie ID {workforce_id}")
    flash("Saisie d’effectif supprimée.")
    return redirect(url_for("workforce"))


@app.route("/export_workforce_excel", methods=["GET"])
def export_workforce_excel():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    clauses=[]; params=[]
    if start_date: clauses.append("w.work_date>=?"); params.append(start_date)
    if end_date: clauses.append("w.work_date<=?"); params.append(end_date)
    scope_where, scope_params = workforce_scope("w")
    if scope_where:
        clauses.append(scope_where.replace(" WHERE ","").strip()); params += scope_params
    where_sql = " WHERE " + " AND ".join(clauses) if clauses else ""
    rows = conn.execute(f"""SELECT w.*, p.name project_name, s.name st_name
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id JOIN subcontractors s ON s.id=w.subcontractor_id
        {where_sql} ORDER BY w.work_date, p.name, s.name""", params).fetchall()
    by_day = conn.execute(f"""SELECT w.work_date, COALESCE(p.name,'') project_name, SUM(w.total) total
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id {where_sql}
        GROUP BY w.work_date, project_name ORDER BY w.work_date""", params).fetchall()
    by_month = conn.execute(f"""SELECT substr(w.work_date,1,7) month, COALESCE(p.name,'') project_name, SUM(w.total) total
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id {where_sql}
        GROUP BY month, project_name ORDER BY month""", params).fetchall()
    by_year = conn.execute(f"""SELECT substr(w.work_date,1,4) year, COALESCE(p.name,'') project_name, SUM(w.total) total
        FROM workforce w LEFT JOIN projects p ON p.id=w.project_id {where_sql}
        GROUP BY year, project_name ORDER BY year""", params).fetchall()
    conn.close()
    wb=Workbook(); ws=wb.active; ws.title="Détails effectifs"
    ws.append(["Date","Projet/Site","Sous-traitant","Hommes","Femmes","Total","Observations","Saisi par","Date saisie"])
    for r in rows: ws.append([r['work_date'],r['project_name'],r['st_name'],r['men'],r['women'],r['total'],r['observations'],r['created_by'],r['created_at']])
    ws2=wb.create_sheet("Totaux journaliers"); ws2.append(["Date","Site","Total"])
    for r in by_day: ws2.append([r['work_date'],r['project_name'],r['total']])
    ws3=wb.create_sheet("Totaux mensuels"); ws3.append(["Mois","Site","Total"])
    for r in by_month: ws3.append([r['month'],r['project_name'],r['total']])
    ws4=wb.create_sheet("Totaux annuels"); ws4.append(["Année","Site","Total"])
    for r in by_year: ws4.append([r['year'],r['project_name'],r['total']])
    fill = PatternFill("solid", fgColor="1F3347")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(len(str(c.value or "")) for c in col)+2, 45)
    filename = EXPORT_DIR / f"Rapport_Effectifs_{start_date or 'debut'}_{end_date or 'fin'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename); log_action("Export Excel effectifs", filename.name)
    return send_file(filename, as_attachment=True)

@app.route("/alerts")
def alerts():
    if not logged() or is_subcontractor(): return redirect(url_for("dashboard"))
    trs = ""
    for level, typ, product, action in get_alert_rows():
        cls = "danger" if level == "CRITIQUE" else "warn"
        trs += f"<tr><td class='{cls}'>{level}</td><td>{typ}</td><td>{product}</td><td>{action}</td></tr>"
    if not trs:
        trs = "<tr><td colspan='4' class='ok'>Aucune alerte majeure détectée.</td></tr>"
    return layout("Alertes QHSE", f"""<h2>Alertes QHSE</h2><div class="help">Alertes automatiques : FDS, produits non reconnus et seuils de stock.</div>
    <table><tr><th>Niveau</th><th>Type</th><th>Produit</th><th>Action recommandée</th></tr>{trs}</table>""")

@app.route("/consolidation")
def consolidation():
    if not logged() or is_subcontractor(): return redirect(url_for("dashboard"))
    trs = "".join([f"<tr><td>{r['projects'] or ''}</td><td>{r['product_name']}</td><td>{r['family'] or ''}</td><td>{r['conditionnement'] or ''}</td><td>{r['utilisation'] or ''}</td><td>{r['fds'] or ''}</td><td>{r['fds_expiry'] or ''}</td><td>{r['pictogrammes'] or ''}</td><td>{r['incompatibilites'] or ''}</td><td>{r['total_in']}</td><td>{r['total_used']}</td><td>{r['total_stock']}</td><td>{r['unit']}</td><td>{r['nb_st']}</td><td>{r['sts']}</td><td>{r['locations'] or ''}</td></tr>" for r in consolidation_rows()])
    export_button = f"<p><a class='btn' href='{url_for('export_excel')}'>Exporter en Excel</a></p>" if can_access_admin_features() else ""
    return layout("Consolidation", f"""<h2>Consolidation générale</h2>{export_button}
    <table><tr><th>Projet</th><th>Produit</th><th>Famille</th><th>Conditionnement</th><th>Utilisation</th><th>FDS</th><th>Exp. FDS</th><th>Pictogrammes</th><th>Incompatibilités</th><th>Entrée</th><th>Utilisée</th><th>Stock</th><th>Unité</th><th>Nb ST</th><th>Sous-traitants</th><th>Lieux stockage</th></tr>{trs}</table>""")

@app.route("/export_excel", methods=["GET","POST"])
def export_excel():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "GET" and not request.args.get("generate"):
        if is_super_admin():
            projects = conn.execute("SELECT * FROM projects ORDER BY name").fetchall()
        else:
            projects = conn.execute("SELECT * FROM projects WHERE id=?", (current_project_id(),)).fetchall()
        sts = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall(); conn.close()
        project_opts = "<option value=''>Tous les projets autorisés</option>" + "".join([f"<option value='{p['id']}'>{p['name']}</option>" for p in projects])
        st_opts = "<option value=''>Tous les sous-traitants</option>" + "".join([f"<option value='{st['id']}'>{st['name']}</option>" for st in sts])
        return layout("Rapport Excel", f"""<h2>Générer rapport Excel</h2>
        <div class="help">Le rapport Excel est réservé au Super Admin et aux Admin Projet. Sélectionne une période, puis filtre éventuellement par projet ou sous-traitant.</div>
        <div class="box"><form method="get">
        <input type="hidden" name="generate" value="1">
        <div class="row"><div><label>Date début</label><input type="date" name="start_date" required placeholder="Début de la période"></div><div><label>Date fin</label><input type="date" name="end_date" required placeholder="Fin de la période"></div></div>
        <div class="row"><div><label>Projet</label><select name="project_id">{project_opts}</select></div><div><label>Sous-traitant</label><select name="subcontractor_id">{st_opts}</select></div></div>
        <button>Générer le rapport Excel</button></form></div>""")
    start_date = request.values.get("start_date", "")
    end_date = request.values.get("end_date", "")
    project_filter = request.values.get("project_id") or None
    st_filter = request.values.get("subcontractor_id") or None
    clauses=[]; params=[]
    if start_date: clauses.append("e.entry_date>=?"); params.append(start_date)
    if end_date: clauses.append("e.entry_date<=?"); params.append(end_date)
    if project_filter: clauses.append("e.project_id=?"); params.append(project_filter)
    if st_filter: clauses.append("e.subcontractor_id=?"); params.append(st_filter)
    if is_project_admin(): clauses.append("e.project_id=?"); params.append(current_project_id())
    where = " WHERE " + " AND ".join(clauses) if clauses else ""

    detail_rows = conn.execute(f"""SELECT e.*, COALESCE(p.root_name,e.declared_product) product_name, p.family,p.conditionnement,COALESCE(e.utilisation,p.utilisation) utilisation,p.fds,p.fds_expiry,p.pictogrammes,p.incompatibilites,p.danger_class,p.epi,s.name st_name,pr.name project_name
        FROM entries e
        LEFT JOIN products p ON p.id=e.product_id
        JOIN subcontractors s ON s.id=e.subcontractor_id
        LEFT JOIN projects pr ON pr.id=e.project_id
        {where}
        ORDER BY e.entry_date, pr.name, s.name, product_name""", params).fetchall()
    summary_rows = conn.execute(f"""SELECT COALESCE(pr.name,'') project_name, COALESCE(p.root_name,e.declared_product) product_name, COALESCE(p.family,'') family, COALESCE(e.utilisation,p.utilisation,'') utilisation, e.unit,
        SUM(e.qty_in) total_in, SUM(e.qty_used) total_used, SUM(e.qty_stock) total_stock,
        COUNT(DISTINCT e.subcontractor_id) nb_st, GROUP_CONCAT(DISTINCT s.name) sts, GROUP_CONCAT(DISTINCT e.storage_location) locations,
        p.fds, p.fds_expiry, p.pictogrammes, p.incompatibilites, p.danger_class, p.epi, p.stock_min, p.stock_max
        FROM entries e
        LEFT JOIN products p ON p.id=e.product_id
        JOIN subcontractors s ON s.id=e.subcontractor_id
        LEFT JOIN projects pr ON pr.id=e.project_id
        {where}
        GROUP BY project_name, product_name, utilisation, e.unit
        ORDER BY project_name, product_name""", params).fetchall()
    conn.close()

    wb = Workbook(); ws = wb.active; ws.title = "Synthèse période"
    ws.append(["Période", f"{start_date or 'début'} au {end_date or 'fin'}"])
    headers = ["Projet","Produit","Famille","Utilisation","FDS","Expiration FDS","Pictogrammes","Incompatibilités","Classe danger","EPI","Entrée totale","Utilisée totale","Stock total","Unité","Stock min","Stock max","Nb sous-traitants","Sous-traitants","Lieux stockage"]
    ws.append(headers)
    for r in summary_rows:
        ws.append([r["project_name"],r["product_name"],r["family"],r["utilisation"],r["fds"],r["fds_expiry"],r["pictogrammes"],r["incompatibilites"],r["danger_class"],r["epi"],r["total_in"],r["total_used"],r["total_stock"],r["unit"],r["stock_min"],r["stock_max"],r["nb_st"],r["sts"],r["locations"]])
    ws_detail = wb.create_sheet("Détails saisies")
    ws_detail.append(["Date","Mois","Projet","Sous-traitant","Produit déclaré","Produit racine","Famille","Utilisation","Entrée","Utilisée","Stock","Unité","FDS saisie","Lieu stockage","Observations"])
    for r in detail_rows:
        ws_detail.append([r["entry_date"],r["month"],r["project_name"],r["st_name"],r["declared_product"],r["product_name"],r["family"],r["utilisation"],r["qty_in"],r["qty_used"],r["qty_stock"],r["unit"],r["fds_available"],r["storage_location"],r["observations"]])
    ws_alert = wb.create_sheet("Alertes")
    ws_alert.append(["Niveau","Type","Produit","Action recommandée"])
    for a in get_alert_rows(): ws_alert.append(list(a))
    ws_about = wb.create_sheet("À propos")
    for row in [["Logiciel",APP_NAME],["Version",VERSION],["Auteur","Kodjotse Eli ADIGBLI"],["Copyright",COPYRIGHT],["Période",f"{start_date or 'début'} au {end_date or 'fin'}"],["Généré le",datetime.now().strftime("%d/%m/%Y %H:%M")]]: ws_about.append(row)
    fill = PatternFill("solid", fgColor="1F3347")
    for sheet in wb.worksheets:
        if sheet.max_row:
            header_row = 2 if sheet.title == "Synthèse période" else 1
            for cell in sheet[header_row]:
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in sheet.columns:
                sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(len(str(c.value or "")) for c in col)+2, 55)
    filename = EXPORT_DIR / f"Rapport_Produits_Chimiques_{start_date or 'debut'}_{end_date or 'fin'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename); log_action("Export Excel période", filename.name)
    return send_file(filename, as_attachment=True)


@app.route("/backups", methods=["GET", "POST"])
def backups():
    if not logged() or not can_access_admin_features():
        return redirect(url_for("dashboard"))
    if not is_super_admin():
        flash("Seul le super admin peut restaurer la base de données.")

    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            path = backup_before_sensitive_operation("manuel")
            flash(f"Sauvegarde créée : {path.name}" if path else "Aucune base disponible à sauvegarder.")
            return redirect(url_for("backups"))

        if action == "restore":
            if not is_super_admin():
                return redirect(url_for("backups"))
            backup_path = validate_backup_filename(request.form.get("filename"))
            confirm = request.form.get("confirm") == "RESTAURER"
            if not backup_path:
                flash("Fichier de sauvegarde invalide.")
                return redirect(url_for("backups"))
            if not confirm:
                flash("Pour restaurer, saisir exactement : RESTAURER")
                return redirect(url_for("backups"))
            before = backup_before_sensitive_operation("avant_restauration")
            shutil.copy2(backup_path, DB_PATH)
            log_action("Restauration base", f"Restauré depuis {backup_path.name}; sauvegarde avant restauration: {before.name if before else 'non créée'}")
            flash("Base restaurée avec succès. Déconnecte-toi puis reconnecte-toi si nécessaire.")
            return redirect(url_for("backups"))

    files = []
    for f in sorted(BACKUP_DIR.glob("qhse_chemical_register_*.db"), key=lambda x: x.stat().st_mtime, reverse=True):
        files.append({
            "name": f.name,
            "date": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "size": round(f.stat().st_size / 1024, 1)
        })
    rows = "".join([f"""<tr><td>{x['date']}</td><td>{x['name']}</td><td>{x['size']} Ko</td><td class='actions'>
        <a class='btn mini' href='{url_for('download_backup', filename=x['name'])}'>Télécharger</a>
        {"<form method='post' style='display:inline-block;margin-left:6px'><input type='hidden' name='action' value='restore'><input type='hidden' name='filename' value='" + x['name'] + "'><input name='confirm' placeholder='RESTAURER' style='width:120px;margin:0'><button class='mini btn-danger'>Restaurer</button></form>" if is_super_admin() else ""}
        </td></tr>""" for x in files]) or "<tr><td colspan='4'>Aucune sauvegarde disponible.</td></tr>"
    return layout("Sauvegardes", f"""<h2>Sauvegarde / restauration</h2>
    <div class='help'><strong>Protection activée :</strong> sauvegarde automatique quotidienne, sauvegarde avant démarrage, et sauvegarde automatique avant toute restauration.</div>
    <div class='box'>
    <form method='post'><input type='hidden' name='action' value='create'><button>Créer une sauvegarde maintenant</button></form>
    <p class='muted'>Les sauvegardes sont stockées dans le dossier <strong>backups</strong>. Pour éviter une perte totale, copie régulièrement ce dossier sur clé USB, disque externe ou cloud.</p>
    </div>
    <table><tr><th>Date</th><th>Fichier</th><th>Taille</th><th>Actions</th></tr>{rows}</table>""")

@app.route("/backups/download/<filename>")
def download_backup(filename):
    if not logged() or not can_access_admin_features():
        return redirect(url_for("dashboard"))
    path = validate_backup_filename(filename)
    if not path:
        flash("Fichier de sauvegarde introuvable.")
        return redirect(url_for("backups"))
    return send_file(path, as_attachment=True)

@app.route("/audit_trail")
def audit_trail():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    page, per_page, offset = pagination_data(20)
    total_logs = conn.execute("SELECT COUNT(*) c FROM audit_log").fetchone()["c"]
    rows = conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset)).fetchall()
    conn.close()
    trs = "".join([f"<tr><td>{r['action_date']}</td><td>{r['username']}</td><td>{r['action']}</td><td>{r['details']}</td></tr>" for r in rows])
    pagination = pagination_html(page, per_page, total_logs)
    return layout("Journal des actions", f"""<h2>Journal des actions</h2><table><tr><th>Date</th><th>Utilisateur</th><th>Action</th><th>Détails</th></tr>{trs}</table>{pagination}""")


@app.route("/theme/<theme_name>")
def set_theme(theme_name):
    if theme_name in {"blue", "green", "sand"}:
        session["theme"] = theme_name
    return redirect(request.referrer or url_for("dashboard"))


def daily_report_scope(prefix="r"):
    if is_super_admin(): return "", []
    if is_project_admin() or is_reader(): return f" WHERE {prefix}.project_id=? ", [current_project_id()]
    if is_subcontractor(): return f" WHERE {prefix}.subcontractor_id=? ", [current_subcontractor_id()]
    return " WHERE 1=0 ", []


def can_manage_daily_report(row):
    if is_super_admin(): return True
    if is_project_admin(): return str(row["project_id"]) == str(current_project_id())
    if is_subcontractor(): return str(row["subcontractor_id"]) == str(current_subcontractor_id())
    return False


@app.route("/daily-reports", methods=["GET", "POST"])
def daily_reports():
    if not logged(): return redirect(url_for("login"))
    conn=db()
    if request.method == "POST":
        if is_reader(): conn.close(); flash("Compte en lecture seule."); return redirect(url_for("daily_reports"))
        project_id=request.form.get("project_id")
        subcontractor_id=request.form.get("subcontractor_id")
        if is_project_admin(): project_id=current_project_id()
        if is_subcontractor():
            subcontractor_id=current_subcontractor_id(); project_id=current_project_id() or request.form.get("project_id")
        if not project_id or not subcontractor_id:
            conn.close(); flash("Projet et sous-traitant obligatoires."); return redirect(url_for("daily_reports"))
        cur=conn.execute("""INSERT INTO daily_reports(report_date,project_id,subcontractor_id,status,general_observations,created_by,created_at,updated_at)
                            VALUES(?,?,?,?,?,?,?,?)""",(request.form.get("report_date") or date.today().isoformat(),project_id,subcontractor_id,"Brouillon","",session.get("username"),datetime.now().strftime("%Y-%m-%d %H:%M:%S"),datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); rid=cur.lastrowid; conn.close(); log_action("Création rapport journalier", f"Rapport ID {rid}"); return redirect(url_for("edit_daily_report", report_id=rid))
    projects=conn.execute("SELECT * FROM projects WHERE status='Actif' ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?",(current_project_id(),)).fetchall()
    sts=conn.execute("SELECT * FROM subcontractors WHERE status='Actif' ORDER BY name").fetchall() if not is_subcontractor() else conn.execute("SELECT * FROM subcontractors WHERE id=?",(current_subcontractor_id(),)).fetchall()
    start=request.args.get("start_date",""); end=request.args.get("end_date",""); st_filter=request.args.get("subcontractor_id",""); project_filter=request.args.get("project_id","")
    scope,params=daily_report_scope("r"); clauses=[]
    if scope: clauses.append(scope.replace(" WHERE ","").strip())
    if start: clauses.append("r.report_date>=?"); params.append(start)
    if end: clauses.append("r.report_date<=?"); params.append(end)
    if st_filter: clauses.append("r.subcontractor_id=?"); params.append(st_filter)
    if project_filter: clauses.append("r.project_id=?"); params.append(project_filter)
    where=" WHERE "+" AND ".join(clauses) if clauses else ""
    page,per_page,offset=pagination_data(); total=conn.execute(f"SELECT COUNT(*) c FROM daily_reports r {where}",params).fetchone()["c"]
    rows=conn.execute(f"""SELECT r.*,p.name project_name,s.name st_name FROM daily_reports r JOIN projects p ON p.id=r.project_id JOIN subcontractors s ON s.id=r.subcontractor_id {where} ORDER BY r.report_date DESC,r.id DESC LIMIT ? OFFSET ?""",params+[per_page,offset]).fetchall()
    conn.close()
    po="".join([f"<option value='{x['id']}'>{x['name']}</option>" for x in projects]); so="".join([f"<option value='{x['id']}'>{x['name']}</option>" for x in sts])
    create="" if is_reader() else f"""<div class='box'><form method='post'><div class='row3'><div><label>Date</label><input type='date' name='report_date' value='{date.today().isoformat()}' required></div><div><label>Projet</label><select name='project_id'>{po}</select></div><div><label>Sous-traitant</label><select name='subcontractor_id'>{so}</select></div></div><button>Créer le rapport journalier</button></form></div>"""
    trs=""
    for r in rows:
        actions=f"<a class='btn mini' href='{url_for('edit_daily_report',report_id=r['id'])}'>{'Modifier' if can_manage_daily_report(r) and not is_reader() else 'Consulter'}</a>"
        if can_manage_daily_report(r) and not is_reader(): actions+=f" <form method='post' action='{url_for('delete_daily_report',report_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer ce rapport ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>"
        trs+=f"<tr><td>{r['report_date']}</td><td>{r['project_name']}</td><td>{r['st_name']}</td><td>{r['status']}</td><td>{r['updated_at'] or ''}</td><td>{actions}</td></tr>"
    trs=trs or "<tr><td colspan='6'>Aucun rapport.</td></tr>"
    admin_buttons=f"<a class='btn' href='{url_for('daily_report_template')}'>Configurer les lignes et colonnes</a> <a class='btn btn-warn' href='{url_for('export_daily_reports_excel')}'>Export Excel / statistiques</a>" if can_access_admin_features() else ""
    filt=f"""<div class='box'><form method='get'><div class='row3'><div><label>Du</label><input type='date' name='start_date' value='{start}'></div><div><label>Au</label><input type='date' name='end_date' value='{end}'></div><div><label>Sous-traitant</label><select name='subcontractor_id'><option value=''>Tous</option>{so}</select></div></div><button>Filtrer</button></form></div>"""
    return layout("Rapports journaliers",f"<h2>Rapport HSE journalier sous-traitant</h2><div class='help'>20 lignes par page. Le modèle est modifiable par les administrateurs : ajout, modification et suppression de lignes ou colonnes.</div>{admin_buttons}<br><br>{create}{filt}<table><tr><th>Date</th><th>Projet</th><th>Sous-traitant</th><th>Statut</th><th>Mise à jour</th><th>Actions</th></tr>{trs}</table>{pagination_html(page,per_page,total)}")


@app.route("/daily-reports/<int:report_id>", methods=["GET","POST"])
def edit_daily_report(report_id):
    if not logged(): return redirect(url_for("login"))
    conn=db(); report=conn.execute("SELECT r.*,p.name project_name,s.name st_name FROM daily_reports r JOIN projects p ON p.id=r.project_id JOIN subcontractors s ON s.id=r.subcontractor_id WHERE r.id=?",(report_id,)).fetchone()
    if not report: conn.close(); flash("Rapport introuvable."); return redirect(url_for("daily_reports"))
    can_edit=can_manage_daily_report(report) and not is_reader()
    if request.method=="POST":
        if not can_edit: conn.close(); flash("Modification non autorisée."); return redirect(url_for("daily_reports"))
        fields=conn.execute("SELECT id FROM daily_report_fields WHERE active=1").fetchall(); cols=conn.execute("SELECT id FROM daily_report_columns WHERE active=1").fetchall()
        for f in fields:
            for c in cols:
                value=request.form.get(f"v_{f['id']}_{c['id']}","")
                conn.execute("""INSERT INTO daily_report_values(report_id,field_id,column_id,value) VALUES(?,?,?,?) ON CONFLICT(report_id,field_id,column_id) DO UPDATE SET value=excluded.value""",(report_id,f['id'],c['id'],value))
        conn.execute("UPDATE daily_reports SET status=?,general_observations=?,updated_at=? WHERE id=?",(request.form.get("status","Brouillon"),request.form.get("general_observations",""),datetime.now().strftime("%Y-%m-%d %H:%M:%S"),report_id)); conn.commit(); log_action("Modification rapport journalier",f"Rapport ID {report_id}"); flash("Rapport enregistré.")
    fields=conn.execute("SELECT * FROM daily_report_fields WHERE active=1 ORDER BY display_order,id").fetchall(); cols=conn.execute("SELECT * FROM daily_report_columns WHERE active=1 ORDER BY display_order,id").fetchall(); values=conn.execute("SELECT field_id,column_id,value FROM daily_report_values WHERE report_id=?",(report_id,)).fetchall(); conn.close()
    vm={(x['field_id'],x['column_id']):x['value'] for x in values}; rows=""; current=None
    for f in fields:
        if f['category']!=current:
            current=f['category']; rows+=f"<tr><th colspan='{3+len(cols)}' style='background:#dcead5;color:#17351f'>{escape(current)}</th></tr>"
        cells="".join([f"<td><input name='v_{f['id']}_{c['id']}' value='{escape(vm.get((f['id'],c['id']),''))}' {'readonly' if not can_edit else ''}></td>" for c in cols])
        rows+=f"<tr><td>{f['display_order']}</td><td>{escape(f['label'])}</td><td>{escape(f['unit'] or '')}</td>{cells}</tr>"
    headers="".join([f"<th>{escape(c['name'])}</th>" for c in cols]); disabled="disabled" if not can_edit else ""
    return layout("Rapport journalier",f"""<h2>Rapport HSE journalier – {report['st_name']}</h2><div class='help'><strong>Date :</strong> {report['report_date']} | <strong>Projet :</strong> {report['project_name']}</div><form method='post'><div class='table-wrap'><table><thead><tr><th>N°</th><th>Indicateur</th><th>Unité</th>{headers}</tr></thead><tbody>{rows}</tbody></table></div><label>Observations générales</label><textarea name='general_observations' { 'readonly' if not can_edit else ''}>{escape(report['general_observations'] or '')}</textarea><label>Statut</label><select name='status' {disabled}><option {'selected' if report['status']=='Brouillon' else ''}>Brouillon</option><option {'selected' if report['status']=='Validé' else ''}>Validé</option></select>{'<button>Enregistrer le rapport</button>' if can_edit else ''} <a class='btn' href='{url_for('daily_reports')}'>Retour</a></form>""")


@app.route("/daily-reports/<int:report_id>/delete",methods=["POST"])
def delete_daily_report(report_id):
    if not logged() or is_reader(): return redirect(url_for("dashboard"))
    conn=db(); row=conn.execute("SELECT * FROM daily_reports WHERE id=?",(report_id,)).fetchone()
    if not row or not can_manage_daily_report(row): conn.close(); flash("Suppression non autorisée."); return redirect(url_for("daily_reports"))
    conn.execute("DELETE FROM daily_report_values WHERE report_id=?",(report_id,)); conn.execute("DELETE FROM daily_reports WHERE id=?",(report_id,)); conn.commit(); conn.close(); log_action("Suppression rapport journalier",f"Rapport ID {report_id}"); flash("Rapport supprimé."); return redirect(url_for("daily_reports"))


@app.route("/daily-report-template",methods=["GET","POST"])
def daily_report_template():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="add_field": conn.execute("INSERT INTO daily_report_fields(category,label,unit,display_order,active) VALUES(?,?,?,?,1)",(request.form.get("category","AUTRES").strip().upper(),request.form.get("label","").strip(),request.form.get("unit",""),safe_int(request.form.get("display_order"),999)))
        elif action=="edit_field": conn.execute("UPDATE daily_report_fields SET category=?,label=?,unit=?,display_order=? WHERE id=?",(request.form.get("category","").strip().upper(),request.form.get("label","").strip(),request.form.get("unit",""),safe_int(request.form.get("display_order"),999),request.form.get("id")))
        elif action=="delete_field": conn.execute("UPDATE daily_report_fields SET active=0 WHERE id=?",(request.form.get("id"),))
        elif action=="add_column": conn.execute("INSERT INTO daily_report_columns(name,display_order,active) VALUES(?,?,1)",(request.form.get("name","").strip(),safe_int(request.form.get("display_order"),999)))
        elif action=="edit_column": conn.execute("UPDATE daily_report_columns SET name=?,display_order=? WHERE id=?",(request.form.get("name","").strip(),safe_int(request.form.get("display_order"),999),request.form.get("id")))
        elif action=="delete_column": conn.execute("UPDATE daily_report_columns SET active=0 WHERE id=?",(request.form.get("id"),))
        conn.commit(); flash("Modèle mis à jour.")
    fields=conn.execute("SELECT * FROM daily_report_fields WHERE active=1 ORDER BY display_order,id").fetchall(); cols=conn.execute("SELECT * FROM daily_report_columns WHERE active=1 ORDER BY display_order,id").fetchall(); conn.close()
    fr="".join([f"<tr><form method='post'><input type='hidden' name='id' value='{f['id']}'><td><input name='display_order' value='{f['display_order']}'></td><td><input name='category' value='{escape(f['category'])}'></td><td><input name='label' value='{escape(f['label'])}'></td><td><input name='unit' value='{escape(f['unit'] or '')}'></td><td><button name='action' value='edit_field' class='mini'>Modifier</button> <button name='action' value='delete_field' class='mini btn-danger'>Supprimer</button></td></form></tr>" for f in fields]); cr="".join([f"<tr><form method='post'><input type='hidden' name='id' value='{c['id']}'><td><input name='display_order' value='{c['display_order']}'></td><td><input name='name' value='{escape(c['name'])}'></td><td><button name='action' value='edit_column' class='mini'>Modifier</button> <button name='action' value='delete_column' class='mini btn-danger'>Supprimer</button></td></form></tr>" for c in cols])
    return layout("Modèle rapport",f"""<h2>Configuration du rapport journalier</h2><div class='help'>Les changements s’appliquent aux prochains affichages. Les anciennes valeurs restent conservées en base.</div><div class='row'><div class='box'><h3>Ajouter une ligne</h3><form method='post'><input type='hidden' name='action' value='add_field'><label>Ordre</label><input name='display_order' value='999'><label>Rubrique</label><input name='category' required><label>Indicateur</label><input name='label' required><label>Unité</label><input name='unit'><button>Ajouter la ligne</button></form></div><div class='box'><h3>Ajouter une colonne</h3><form method='post'><input type='hidden' name='action' value='add_column'><label>Ordre</label><input name='display_order' value='99'><label>Nom de la colonne</label><input name='name' required><button>Ajouter la colonne</button></form></div></div><h3>Lignes</h3><table><tr><th>Ordre</th><th>Rubrique</th><th>Indicateur</th><th>Unité</th><th>Actions</th></tr>{fr}</table><h3>Colonnes</h3><table><tr><th>Ordre</th><th>Nom</th><th>Actions</th></tr>{cr}</table>""")


@app.route("/daily-reports/export",methods=["GET"])
def export_daily_reports_excel():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    if not request.args.get("generate"):
        conn=db(); projects=conn.execute("SELECT * FROM projects ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?",(current_project_id(),)).fetchall(); sts=conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall(); conn.close(); po="<option value=''>Tous</option>"+"".join([f"<option value='{x['id']}'>{x['name']}</option>" for x in projects]); so="<option value=''>Tous</option>"+"".join([f"<option value='{x['id']}'>{x['name']}</option>" for x in sts]); return layout("Export rapports",f"""<h2>Export Excel et statistiques</h2><div class='box'><form method='get'><input type='hidden' name='generate' value='1'><div class='row3'><div><label>Du</label><input type='date' name='start_date' required></div><div><label>Au</label><input type='date' name='end_date' required></div><div><label>Période statistique</label><select name='period'><option value='day'>Jour</option><option value='week'>Semaine</option><option value='month'>Mois</option></select></div></div><div class='row3'><div><label>Projet</label><select name='project_id'>{po}</select></div><div><label>Sous-traitant</label><select name='subcontractor_id'>{so}</select></div><div><label>Format</label><input value='Couleur standard QHSE Manager Pro' readonly></div></div><button>Générer l’Excel</button></form></div>""")
    start=request.args.get("start_date"); end=request.args.get("end_date"); period=request.args.get("period","day"); project=request.args.get("project_id"); st=request.args.get("subcontractor_id"); palette='blue'
    conn=db(); clauses=["r.report_date>=?","r.report_date<=?"]; params=[start,end]
    if is_project_admin(): clauses.append("r.project_id=?"); params.append(current_project_id())
    elif project: clauses.append("r.project_id=?"); params.append(project)
    if st: clauses.append("r.subcontractor_id=?"); params.append(st)
    where=" WHERE "+" AND ".join(clauses)
    rows=conn.execute(f"""SELECT r.report_date,p.name project_name,s.name st_name,f.category,f.label,f.unit,c.name column_name,v.value FROM daily_reports r JOIN projects p ON p.id=r.project_id JOIN subcontractors s ON s.id=r.subcontractor_id JOIN daily_report_values v ON v.report_id=r.id JOIN daily_report_fields f ON f.id=v.field_id JOIN daily_report_columns c ON c.id=v.column_id {where} ORDER BY r.report_date,s.name,f.display_order,c.display_order""",params).fetchall(); conn.close()
    colors={"blue":("1F4E78","D9EAF7"),"green":("217346","E2F0D9"),"sand":("806000","FFF2CC")}; dark,light=colors.get(palette,colors['blue']); wb=Workbook(); ws=wb.active; ws.title="Détails journaliers"; ws.append(["Date","Projet","Sous-traitant","Rubrique","Indicateur","Unité","Colonne","Valeur"])
    for r in rows: ws.append([r['report_date'],r['project_name'],r['st_name'],r['category'],r['label'],r['unit'],r['column_name'],r['value']])
    stat=wb.create_sheet("Statistiques"); stat.append(["Période","Sous-traitant","Indicateur","Unité","Somme numérique","Valeurs renseignées"]); agg={}
    for r in rows:
        d=datetime.strptime(r['report_date'],"%Y-%m-%d").date(); keyp=d.isoformat() if period=='day' else (f"{d.isocalendar().year}-S{d.isocalendar().week:02d}" if period=='week' else d.strftime('%Y-%m'))
        k=(keyp,r['st_name'],r['label'],r['unit']); a=agg.setdefault(k,[0.0,0]); val=(r['value'] or '').strip()
        if val:
            a[1]+=1
            try: a[0]+=float(val.replace(',','.'))
            except: pass
    for k,a in sorted(agg.items()): stat.append(list(k)+[a[0],a[1]])
    for sheet in wb.worksheets:
        for cell in sheet[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=dark); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            if row[0].row%2==0:
                for cell in row: cell.fill=PatternFill("solid",fgColor=light)
        sheet.freeze_panes="A2"; sheet.auto_filter.ref=sheet.dimensions
        for col in sheet.columns: sheet.column_dimensions[get_column_letter(col[0].column)].width=min(max(len(str(c.value or '')) for c in col)+2,55)
    filename=EXPORT_DIR/f"Rapport_HSE_Journalier_{start}_{end}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"; wb.save(filename); log_action("Export rapports HSE journaliers",filename.name); return send_file(filename,as_attachment=True)

# ---------- Plan d'action mensuel ----------
def action_plan_scope_sql(alias="a"):
    if is_super_admin():
        return "", []
    if is_project_admin() or is_reader():
        return f" WHERE {alias}.project_id=? ", [current_project_id()]
    return " WHERE 1=0 ", []

def can_manage_action_plan(row):
    if is_super_admin(): return True
    if is_project_admin(): return str(row["project_id"]) == str(current_project_id())
    return False

def css_token(value):
    return normalize(value).lower().replace(" ", "").replace("é", "e").replace("è", "e").replace("ê", "e")

@app.route("/action-plans", methods=["GET", "POST"])
def action_plans():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn = db()
    if request.method == "POST":
        project_id = request.form.get("project_id") or current_project_id()
        if is_project_admin(): project_id = current_project_id()
        conn.execute("""INSERT INTO action_plans(project_id,source,opened_date,risk_description,required_action,responsible,due_date,priority,status,comments,created_by,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", (project_id, request.form.get("source",""), request.form.get("opened_date"), request.form.get("risk_description",""), request.form.get("required_action",""), request.form.get("responsible",""), request.form.get("due_date",""), request.form.get("priority","Moyenne"), request.form.get("status","Ouvert"), request.form.get("comments",""), session.get("username",""), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); log_action("Création point d’action", request.form.get("risk_description","")[:120]); flash("Point d’action ajouté.")
    page = max(1, safe_int(request.args.get("page"), 1)); per_page = 20
    month = request.args.get("month", date.today().strftime("%Y-%m")); status_filter=request.args.get("status",""); project_filter=request.args.get("project_id","")
    clauses=[]; params=[]
    if month: clauses.append("substr(a.opened_date,1,7)=?"); params.append(month)
    if status_filter: clauses.append("a.status=?"); params.append(status_filter)
    if project_filter: clauses.append("a.project_id=?"); params.append(project_filter)
    if is_project_admin(): clauses.append("a.project_id=?"); params.append(current_project_id())
    where=" WHERE "+" AND ".join(clauses) if clauses else ""
    total=conn.execute(f"SELECT COUNT(*) c FROM action_plans a {where}",params).fetchone()["c"]
    rows=conn.execute(f"""SELECT a.*,p.name project_name FROM action_plans a LEFT JOIN projects p ON p.id=a.project_id {where}
        ORDER BY a.opened_date DESC,a.id DESC LIMIT ? OFFSET ?""",params+[per_page,(page-1)*per_page]).fetchall()
    projects=conn.execute("SELECT * FROM projects ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?",(current_project_id(),)).fetchall(); conn.close()
    po="".join([f"<option value='{x['id']}' {'selected' if str(project_filter)==str(x['id']) else ''}>{escape(x['name'])}</option>" for x in projects])
    trs=""
    for r in rows:
        pc="priority-"+css_token(r['priority']); sc="status-"+css_token(r['status'])
        actions=f"<a class='btn mini' href='{url_for('edit_action_plan', action_id=r['id'])}'>Modifier</a> <form method='post' action='{url_for('delete_action_plan', action_id=r['id'])}' style='display:inline' onsubmit='return confirm(&quot;Supprimer ce point d’action ?&quot;)'><button class='mini btn-danger'>Supprimer</button></form>" if can_manage_action_plan(r) else ""
        trs+=f"<tr><td>{escape(r['source'] or '')}</td><td>{r['opened_date']}</td><td>{escape(r['risk_description'])}</td><td>{escape(r['required_action'] or '')}</td><td>{escape(r['responsible'] or '')}</td><td>{r['due_date'] or ''}</td><td class='{pc}'>{r['priority']}</td><td class='{sc}'>{r['status']}</td><td>{escape(r['comments'] or '')}</td><td>{escape(r['project_name'] or '')}</td><td class='actions'>{actions}</td></tr>"
    trs=trs or "<tr><td colspan='11'>Aucun point d’action.</td></tr>"
    pag=pagination_html(page,per_page,total)
    return layout("Plan d’action",f"""<h2>Plan d’action mensuel</h2><div class='help'>La couleur des cellules Priorité et Statut s’applique automatiquement selon le choix effectué.</div>
    <div class='box'><form method='post'><div class='row3'><div><label>Projet</label><select name='project_id'>{po}</select></div><div><label>Source</label><select name='source'><option>Inspection</option><option>Incident/Accident</option><option>Réunion</option><option>Audit</option><option>Autre</option></select></div><div><label>Date d’ouverture</label><input type='date' name='opened_date' value='{date.today().isoformat()}' required></div></div>
    <label>Description du risque</label><textarea name='risk_description' required></textarea><label>Action requise</label><textarea name='required_action'></textarea>
    <div class='row3'><div><label>Responsable</label><input name='responsible'></div><div><label>Date prévue de fermeture</label><input type='date' name='due_date'></div><div><label>Priorité</label><select name='priority'><option>Critique</option><option>Élevée</option><option selected>Moyenne</option><option>Faible</option></select></div></div>
    <div class='row'><div><label>Statut</label><select name='status'><option selected>Ouvert</option><option>En cours</option><option>Fermé</option><option>Suspendu</option></select></div><div><label>Commentaires, remarques, preuves</label><textarea name='comments'></textarea></div></div><button>Ajouter le point d’action</button></form></div>
    <div class='box'><form method='get'><div class='row3'><div><label>Mois</label><input type='month' name='month' value='{month}'></div><div><label>Projet</label><select name='project_id'><option value=''>Tous</option>{po}</select></div><div><label>Statut</label><select name='status'><option value=''>Tous</option><option>Ouvert</option><option>En cours</option><option>Fermé</option><option>Suspendu</option></select></div></div><button>Filtrer</button> <a class='btn' href='{url_for('export_action_plans',month=month,project_id=project_filter,status=status_filter)}'>Exporter Excel</a></form></div>
    <div class='table-wrap'><table><thead><tr><th>Source</th><th>Date d’ouverture</th><th>Description du risque</th><th>Action requise</th><th>Responsable</th><th>Date fermeture</th><th>Priorité</th><th>Statut</th><th>Commentaires</th><th>Projet</th><th>Actions</th></tr></thead><tbody>{trs}</tbody></table></div>{pag}""")

@app.route("/action-plans/<int:action_id>/edit", methods=["GET","POST"])
def edit_action_plan(action_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM action_plans WHERE id=?",(action_id,)).fetchone()
    if not r or not can_manage_action_plan(r): conn.close(); flash("Accès refusé."); return redirect(url_for("action_plans"))
    if request.method=="POST":
        conn.execute("""UPDATE action_plans SET source=?,opened_date=?,risk_description=?,required_action=?,responsible=?,due_date=?,priority=?,status=?,comments=?,updated_at=? WHERE id=?""",(request.form.get("source",""),request.form.get("opened_date"),request.form.get("risk_description",""),request.form.get("required_action",""),request.form.get("responsible",""),request.form.get("due_date",""),request.form.get("priority","Moyenne"),request.form.get("status","Ouvert"),request.form.get("comments",""),datetime.now().strftime("%Y-%m-%d %H:%M:%S"),action_id)); conn.commit(); conn.close(); log_action("Modification point d’action",f"ID {action_id}"); flash("Point d’action modifié."); return redirect(url_for("action_plans"))
    conn.close()
    opts=lambda vals,current: "".join([f"<option {'selected' if current==x else ''}>{x}</option>" for x in vals])
    return layout("Modifier point d’action",f"""<h2>Modifier le point d’action</h2><div class='box'><form method='post'><label>Source</label><select name='source'>{opts(['Inspection','Incident/Accident','Réunion','Audit','Autre'],r['source'])}</select><label>Date d’ouverture</label><input type='date' name='opened_date' value='{r['opened_date']}' required><label>Description du risque</label><textarea name='risk_description' required>{escape(r['risk_description'])}</textarea><label>Action requise</label><textarea name='required_action'>{escape(r['required_action'] or '')}</textarea><div class='row3'><div><label>Responsable</label><input name='responsible' value='{escape(r['responsible'] or '')}'></div><div><label>Date fermeture</label><input type='date' name='due_date' value='{r['due_date'] or ''}'></div><div><label>Priorité</label><select name='priority'>{opts(['Critique','Élevée','Moyenne','Faible'],r['priority'])}</select></div></div><label>Statut</label><select name='status'>{opts(['Ouvert','En cours','Fermé','Suspendu'],r['status'])}</select><label>Commentaires</label><textarea name='comments'>{escape(r['comments'] or '')}</textarea><button>Enregistrer</button> <a class='btn' href='{url_for('action_plans')}'>Annuler</a></form></div>""")

@app.route("/action-plans/<int:action_id>/delete",methods=["POST"])
def delete_action_plan(action_id):
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); r=conn.execute("SELECT * FROM action_plans WHERE id=?",(action_id,)).fetchone()
    if not r or not can_manage_action_plan(r): conn.close(); flash("Suppression non autorisée."); return redirect(url_for("action_plans"))
    conn.execute("DELETE FROM action_plans WHERE id=?",(action_id,)); conn.commit(); conn.close(); log_action("Suppression point d’action",f"ID {action_id}"); flash("Point d’action supprimé."); return redirect(url_for("action_plans"))

@app.route("/action-plans/export")
def export_action_plans():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); clauses=[]; params=[]
    month=request.args.get("month",""); project=request.args.get("project_id",""); status=request.args.get("status","")
    if month: clauses.append("substr(a.opened_date,1,7)=?"); params.append(month)
    if project: clauses.append("a.project_id=?"); params.append(project)
    if status: clauses.append("a.status=?"); params.append(status)
    if is_project_admin(): clauses.append("a.project_id=?"); params.append(current_project_id())
    where=" WHERE "+" AND ".join(clauses) if clauses else ""
    rows=conn.execute(f"SELECT a.*,p.name project_name FROM action_plans a LEFT JOIN projects p ON p.id=a.project_id {where} ORDER BY a.opened_date",params).fetchall(); conn.close()
    wb=Workbook(); ws=wb.active; ws.title="Plan d'action"; ws.append(["Source","Date d'ouverture","Description du risque","Action requise","Responsable","Date fermeture","Priorité","Statut","Commentaires","Projet"])
    fills={"Critique":"FF4D4D","Élevée":"FF0000","Moyenne":"FFC000","Faible":"92D050","Ouvert":"92D050","En cours":"FFC000","Fermé":"19AEE4","Suspendu":"BFBFBF"}
    for r in rows:
        ws.append([r['source'],r['opened_date'],r['risk_description'],r['required_action'],r['responsible'],r['due_date'],r['priority'],r['status'],r['comments'],r['project_name']]); rr=ws.max_row
        ws.cell(rr,7).fill=PatternFill('solid',fgColor=fills.get(r['priority'],'FFFFFF')); ws.cell(rr,8).fill=PatternFill('solid',fgColor=fills.get(r['status'],'FFFFFF'))
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill('solid',fgColor="1F3347"); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(len(str(x.value or '')) for x in col)+2,55)
    fn=EXPORT_DIR/f"Plan_action_{month or 'tous'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"; wb.save(fn); return send_file(fn,as_attachment=True)

# ---------- Messagerie et rappels ----------
def smtp_ready():
    return all(os.environ.get(k) for k in ["SMTP_HOST","SMTP_PORT","SMTP_USERNAME","SMTP_PASSWORD","SMTP_FROM"])

def send_email_message(recipients, subject, body):
    recipients=[x.strip() for x in recipients if x and x.strip()]
    if not recipients: return False,"Aucun destinataire avec adresse email."
    if not smtp_ready(): return False,"SMTP non configuré sur Render."
    msg=EmailMessage(); msg["Subject"]=subject; msg["From"]=os.environ["SMTP_FROM"]; msg["To"]=", ".join(recipients); msg.set_content(body)
    try:
        host=os.environ["SMTP_HOST"]; port=int(os.environ.get("SMTP_PORT","587")); use_ssl=os.environ.get("SMTP_USE_SSL","false").lower()=="true"
        server=smtplib.SMTP_SSL(host,port,timeout=30) if use_ssl else smtplib.SMTP(host,port,timeout=30)
        if not use_ssl: server.starttls()
        server.login(os.environ["SMTP_USERNAME"],os.environ["SMTP_PASSWORD"]); server.send_message(msg); server.quit(); return True,f"Envoyé à {len(recipients)} destinataire(s)."
    except Exception as exc: return False,f"Erreur SMTP : {exc}"

def log_message(channel,subject,message,recipients,project_id=None,subcontractor_id=None,status="Préparé"):
    conn=db(); conn.execute("INSERT INTO message_log(sent_at,sender,channel,subject,message,recipients,project_id,subcontractor_id,delivery_status) VALUES(?,?,?,?,?,?,?,?,?)",(datetime.now().strftime("%Y-%m-%d %H:%M:%S"),session.get("username","system"),channel,subject,message,recipients,project_id,subcontractor_id,status)); conn.commit(); conn.close()

def get_message_recipients(project_id=None,subcontractor_id=None):
    conn=db(); clauses=["status='Actif'","email IS NOT NULL","trim(email)<>''"]; params=[]
    if subcontractor_id: clauses.append("id=?"); params.append(subcontractor_id)
    elif project_id:
        ids=[r['subcontractor_id'] for r in conn.execute("SELECT DISTINCT subcontractor_id FROM entries WHERE project_id=? AND subcontractor_id IS NOT NULL",(project_id,)).fetchall()]
        if ids: clauses.append("id IN (%s)" % ",".join("?"*len(ids))); params+=ids
        else: clauses.append("1=0")
    rows=conn.execute("SELECT * FROM subcontractors WHERE "+" AND ".join(clauses)+" ORDER BY name",params).fetchall(); conn.close(); return rows

@app.route("/messages",methods=["GET","POST"])
def message_center():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    conn=db(); projects=conn.execute("SELECT * FROM projects ORDER BY name").fetchall() if is_super_admin() else conn.execute("SELECT * FROM projects WHERE id=?",(current_project_id(),)).fetchall()
    if is_super_admin():
        sts=conn.execute("SELECT * FROM subcontractors WHERE status='Actif' ORDER BY name").fetchall()
    else:
        sts=conn.execute("""SELECT DISTINCT s.* FROM subcontractors s LEFT JOIN entries e ON e.subcontractor_id=s.id LEFT JOIN users u ON u.subcontractor_id=s.id WHERE s.status='Actif' AND (e.project_id=? OR u.project_id=?) ORDER BY s.name""",(current_project_id(),current_project_id())).fetchall()
    if request.method=="POST":
        action=request.form.get("action"); project_id=request.form.get("project_id") or None; st_id=request.form.get("subcontractor_id") or None; subject=request.form.get("subject","").strip(); message=request.form.get("message","").strip()
        if is_project_admin(): project_id=current_project_id()
        if action=="save_project_contact":
            pid=request.form.get("config_project_id"); conn.execute("UPDATE projects SET whatsapp_group_link=?,notification_email=? WHERE id=?",(request.form.get("whatsapp_group_link",""),request.form.get("notification_email",""),pid)); conn.commit(); flash("Coordonnées du projet enregistrées."); conn.close(); return redirect(url_for("message_center"))
        recipients=get_message_recipients(project_id,st_id)
        emails=[x['email'] for x in recipients if x['email']]
        if action=="email":
            ok,info=send_email_message(emails,subject,message); log_message("Email",subject,message,", ".join(emails),project_id,st_id,"Envoyé" if ok else info); flash(info)
        elif action=="whatsapp":
            target=recipients[0] if len(recipients)==1 else None
            if target and target['phone']:
                phone=''.join(ch for ch in target['phone'] if ch.isdigit()); link=f"https://wa.me/{phone}?text={urllib.parse.quote(message)}"; log_message("WhatsApp",subject,message,target['phone'],project_id,st_id,"Lien généré"); conn.close(); return redirect(link)
            project=conn.execute("SELECT whatsapp_group_link FROM projects WHERE id=?",(project_id,)).fetchone() if project_id else None
            if project and project['whatsapp_group_link']:
                log_message("WhatsApp groupe",subject,message,project['whatsapp_group_link'],project_id,st_id,"Groupe ouvert - message à coller"); flash("Le groupe WhatsApp va s’ouvrir. Copie le message préparé dans le formulaire."); link=project['whatsapp_group_link']; conn.close(); return redirect(link)
            flash("Renseigne le téléphone du sous-traitant ou le lien du groupe WhatsApp du projet.")
    conn.close(); po="".join([f"<option value='{x['id']}'>{escape(x['name'])}</option>" for x in projects]); so="".join([f"<option value='{x['id']}'>{escape(x['name'])} — {escape(x['email'] or 'sans email')}</option>" for x in sts]); config="".join([f"<option value='{x['id']}'>{escape(x['name'])}</option>" for x in projects])
    return layout("Messagerie",f"""<h2>Centre de messages</h2><div class='help'>Les emails partent automatiquement lorsque SMTP est configuré. Pour WhatsApp, l’application ouvre une conversation directe avec message prérempli, ou le groupe du projet. WhatsApp ne permet pas à une application web d’envoyer silencieusement dans un groupe sans API professionnelle.</div>
    <div class='row'><div class='box'><h3>Message personnalisé ou groupé</h3><form method='post'><div class='row'><div><label>Projet</label><select name='project_id'><option value=''>Tous les projets autorisés</option>{po}</select></div><div><label>Sous-traitant précis</label><select name='subcontractor_id'><option value=''>Tous selon le projet</option>{so}</select></div></div><label>Objet</label><input name='subject' required><label>Message</label><textarea name='message' rows='8' required></textarea><button name='action' value='email'>Envoyer par email</button> <button name='action' value='whatsapp' class='btn-warn'>Ouvrir WhatsApp</button></form></div>
    <div class='box'><h3>Coordonnées de notification du projet</h3><form method='post'><input type='hidden' name='action' value='save_project_contact'><label>Projet</label><select name='config_project_id'>{config}</select><label>Email de coordination</label><input name='notification_email' type='email'><label>Lien d’invitation / ouverture du groupe WhatsApp</label><input name='whatsapp_group_link' placeholder='https://chat.whatsapp.com/...'><button>Enregistrer</button></form><p class='muted'>Pour un envoi direct individuel, renseigne le téléphone dans la fiche du sous-traitant.</p></div></div>
    <div class='help'><strong>Rappels automatiques prévus :</strong> produits chimiques le 26 du mois et effectifs à 10h si aucune saisie n’est détectée. Leur exécution nécessite les tâches planifiées décrites dans le guide v2.8.</div>""")

@app.route("/messages/history")
def message_history():
    if not logged() or not can_access_admin_features(): return redirect(url_for("dashboard"))
    page=max(1,safe_int(request.args.get("page"),1)); per=20; conn=db(); total=conn.execute("SELECT COUNT(*) c FROM message_log").fetchone()['c']; rows=conn.execute("SELECT * FROM message_log ORDER BY id DESC LIMIT ? OFFSET ?",(per,(page-1)*per)).fetchall(); conn.close(); trs="".join([f"<tr><td>{r['sent_at']}</td><td>{escape(r['sender'] or '')}</td><td>{escape(r['channel'] or '')}</td><td>{escape(r['subject'] or '')}</td><td>{escape(r['recipients'] or '')}</td><td>{escape(r['delivery_status'] or '')}</td></tr>" for r in rows]) or "<tr><td colspan='6'>Aucun envoi.</td></tr>"; return layout("Historique messages",f"<h2>Historique des messages</h2><table><tr><th>Date</th><th>Expéditeur</th><th>Canal</th><th>Objet</th><th>Destinataires</th><th>Résultat</th></tr>{trs}</table>{pagination_html(page,per,total)}")

def task_authorized():
    token=os.environ.get("TASK_TOKEN",""); supplied=request.headers.get("X-Task-Token") or request.args.get("token",""); return bool(token) and supplied==token

def run_chemical_reminder(force=False):
    if date.today().day!=26 and not force: return {"status":"skipped","reason":"not_day_26"}
    recipients=get_message_recipients(); emails=[x['email'] for x in recipients]; subject=f"Rappel – Rapport produits chimiques {date.today().strftime('%m/%Y')}"; body="Bonjour,\n\nMerci de soumettre et mettre à jour votre rapport mensuel des produits chimiques avant la fin du mois dans QHSE Manager Pro.\n\nCordialement,\nAdministration QHSE"
    ok,info=send_email_message(emails,subject,body); log_message("Email automatique",subject,body,", ".join(emails),status="Envoyé" if ok else info); return {"status":"ok" if ok else "error","detail":info}

def run_workforce_reminder(force=False):
    if datetime.now().hour<10 and not force: return {"status":"skipped","reason":"before_10"}
    conn=db(); today=date.today().isoformat(); pairs=conn.execute("""SELECT DISTINCT u.project_id,u.subcontractor_id,s.email,s.name,p.name project_name FROM users u JOIN subcontractors s ON s.id=u.subcontractor_id LEFT JOIN projects p ON p.id=u.project_id WHERE u.role='subcontractor' AND u.status='Actif' AND s.email IS NOT NULL AND trim(s.email)<>''""").fetchall(); missing=[]
    for x in pairs:
        c=conn.execute("SELECT COUNT(*) c FROM workforce WHERE work_date=? AND project_id=? AND subcontractor_id=?",(today,x['project_id'],x['subcontractor_id'])).fetchone()['c']
        if c==0: missing.append(x)
    conn.close(); sent=0
    for x in missing:
        subject=f"Rappel effectif journalier – {x['project_name'] or 'Projet'}"; body=f"Bonjour {x['name']},\n\nAucune saisie d’effectif n’a été détectée ce jour à 10h. Merci de renseigner l’effectif journalier dans QHSE Manager Pro.\n\nCordialement,\nAdministration QHSE"; ok,info=send_email_message([x['email']],subject,body); log_message("Email automatique",subject,body,x['email'],x['project_id'],x['subcontractor_id'],"Envoyé" if ok else info); sent+=1 if ok else 0
    return {"status":"ok","missing":len(missing),"sent":sent}

@app.route("/tasks/monthly-chemical-reminder",methods=["GET","POST"])
def task_monthly_chemical_reminder():
    if not task_authorized(): return {"status":"forbidden"},403
    return run_chemical_reminder(request.args.get("force")=="1")

@app.route("/tasks/workforce-reminder",methods=["GET","POST"])
def task_workforce_reminder():
    if not task_authorized(): return {"status":"forbidden"},403
    return run_workforce_reminder(request.args.get("force")=="1")


@app.route("/health")
def health():
    return {"status": "ok", "app": APP_NAME, "version": VERSION}

@app.route("/about")
def about():
    if not logged(): return redirect(url_for("login"))
    return layout("À propos", f"""<h2>À propos</h2><div class="box">
    <p><strong>{APP_NAME}</strong></p><p>Version : <strong>{VERSION}</strong></p>
    <p>Développé par : <strong>Kodjotse Eli ADIGBLI</strong></p><p>{COPYRIGHT}</p></div>""")

def open_browser():
    webbrowser.open_new("http://127.0.0.1:5000")

# Cloud initialization
try:
    protect_database_on_startup()
    init_db()
except Exception as cloud_init_error:
    print("Database initialization warning:", cloud_init_error)

if __name__ == "__main__":
    protect_database_on_startup()
    init_db()
    cloud_mode = os.environ.get("CLOUD_MODE", "false").lower() == "true"
    port = int(os.environ.get("PORT", "5000"))
    host = "0.0.0.0" if cloud_mode else "127.0.0.1"
    if not cloud_mode:
        threading.Timer(1.2, open_browser).start()
    app.run(host=host, port=port, debug=False)
